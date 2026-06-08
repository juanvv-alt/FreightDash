import csv
import json
import os
import tempfile
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO

import pdfplumber
from django.contrib import messages
from django.db import DatabaseError, ProgrammingError, transaction
from django.db.models import Max
from django.db.utils import OperationalError
from django.http import HttpResponse, JsonResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
import logging
from .models import RouteParameters
from .models import (
    AvailableIndex,
    ComparisonVessel,
    CustomIndexPreset,
    DailyIndexValue,
    FFACurve,
    FFACurvePeriod,
    FreightVoyage,
    VesselFuelConsumption,
)
from .ffa_utils import parse_ffa_text, resolve_employment_periods
from .forms import TCECalculatorForm
from .calculators import (
    calculate_fuel_and_days,
    calculate_tce,
    calculate_freight_from_tce,
    calculate_vessel_comparison,
)
from bisect import bisect_right


logger = logging.getLogger(__name__)


VESSEL_INDEX_GROUPS = {
    'capesize': {
        'label': 'Capesize',
        'indices': [
            'C2', 'C3', 'C5', 'C7', 'C8_14', 'C9_14', 'C10_14', 'C14', 'C16', 'C17',
            'BCI 5TC', 'BCI 4TC', 'BCI Index', 'Minicape (Old)', 'Minicape (Broker)',
            'Minicape (AVG)', 'Forward Capesize (Current)', 'Forward Capesize (Next MTH)',
            'Forward Panamax (Current)', 'Forward Panamax (Next Mth)',
            'Forward Minicape (Old) (Current)', 'Forward Minicape (Old) (Next Mth)',
            'Forward Minicape (Broker) (Current)', 'Forward Minicape (Broker) (Next Mth)',
        ],
    },
    'panamax': {
        'label': 'Panamax',
        'indices': ['P1A_82', 'P2A_82', 'P3A_82', 'P4A_82', 'P5_82', 'BPI 82TC', 'BPI Index'],
    },
    'supramax': {
        'label': 'Supramax',
        'indices': ['S1B_58', 'S1C_58', 'S2_58', 'S3_58', 'S4A_58', 'S4B_58', 'BSI 58TC', 'BSI Index'],
    },
    'handysize': {
        'label': 'Handysize',
        'indices': ['HS1', 'HS2', 'HS3', 'HS4', 'HS5', 'BHSI 38TC', 'BHSI Index'],
    },
    'bunker': {
        'label': 'Bunker',
        'indices': [
            'Singapore IFO 380', 'Singapore MGO', 'Singapore VLSFO',
            'Hong Kong IFO 380', 'Hong Kong MGO', 'Hong Kong VLSFO',
            'Rotterdam IFO 380', 'Rotterdam MGO', 'Rotterdam VLSFO',
            'Brent', 'WTI',
        ],
    },
}


def _all_known_indices():
    db_indices = list(
        AvailableIndex.objects.filter(is_active=True)
        .order_by('vessel_size', 'order', 'name')
        .values_list('name', flat=True)
    )
    if db_indices:
        return db_indices

    seen = set()
    ordered = []
    for config in VESSEL_INDEX_GROUPS.values():
        for index_name in config['indices']:
            if index_name in seen:
                continue
            seen.add(index_name)
            ordered.append(index_name)
    return ordered


def _indices_for_vessel(vessel_key):
    db_indices = list(
        AvailableIndex.objects.filter(is_active=True, vessel_size=vessel_key)
        .order_by('order', 'name')
        .values_list('name', flat=True)
    )
    if db_indices:
        return db_indices
    return VESSEL_INDEX_GROUPS[vessel_key]['indices']


def _build_rows(start_date, end_date, selected_indices):
    total_days = (end_date - start_date).days + 1
    date_rows = [
        end_date - timedelta(days=offset)
        for offset in range(total_days)
        if (end_date - timedelta(days=offset)).weekday() < 5
    ]
    values_map = {}

    if selected_indices:
        for value in DailyIndexValue.objects.filter(
            index__name__in=selected_indices,
            date__gte=start_date,
            date__lte=end_date,
            index__is_active=True,
        ).select_related('index'):
            values_map[(value.date, value.index.name)] = value.value

    rows = []
    for date_value in date_rows:
        rows.append(
            {
                'date': date_value,
                'values': {
                    index_name: values_map.get((date_value, index_name))
                    for index_name in selected_indices
                },
            }
        )

    return rows, bool(values_map)


def indices_redirect(request):
    return redirect('voyage:indices_by_vessel', vessel='capesize')


def indices_dashboard(request, vessel):
    vessel_key = vessel.lower()
    if vessel_key not in VESSEL_INDEX_GROUPS:
        return redirect('voyage:indices_by_vessel', vessel='capesize')

    vessel_config = VESSEL_INDEX_GROUPS[vessel_key]
    all_indices = _indices_for_vessel(vessel_key)
    today = timezone.localdate()

    default_start = today - timedelta(days=30)
    start_date_raw = request.GET.get('start_date', default_start.isoformat())
    end_date_raw = request.GET.get('end_date', today.isoformat())

    try:
        start_date = timezone.datetime.strptime(start_date_raw, '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start

    try:
        end_date = timezone.datetime.strptime(end_date_raw, '%Y-%m-%d').date()
    except ValueError:
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    if (end_date - start_date).days > 365:
        start_date = end_date - timedelta(days=365)

    selected_indices_raw = request.GET.getlist('indices')
    if 'indices' not in request.GET:
        selected_indices = list(all_indices)
    else:
        selected_indices = []
        seen_indices = set()
        for index_name in selected_indices_raw:
            cleaned = index_name.strip()
            if not cleaned or len(cleaned) > 80:
                continue
            if cleaned in seen_indices:
                continue
            seen_indices.add(cleaned)
            selected_indices.append(cleaned)

    rows, has_data = _build_rows(start_date, end_date, selected_indices)

    # CSV download
    if request.GET.get('download') == '1' and selected_indices:
        response = HttpResponse(content_type='text/csv')
        safe_vessel = vessel_key.replace('/', '_')
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_vessel}_indices_{start_date}_{end_date}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(['Date'] + selected_indices)
        for row in rows:
            writer.writerow(
                [row['date'].strftime('%Y-%m-%d')]
                + [row['values'].get(idx) if row['values'].get(idx) is not None else '' for idx in selected_indices]
            )
        return response

    section_label = 'Bunkers' if vessel_key == 'bunker' else 'Indices'

    context = {
        'vessel_key': vessel_key,
        'vessel_label': vessel_config['label'],
        'vessel_menu': [
            {'key': key, 'label': config['label']}
            for key, config in VESSEL_INDEX_GROUPS.items()
            if key != 'bunker'
        ],
        'all_indices': all_indices,
        'selected_indices': selected_indices,
        'start_date': start_date,
        'end_date': end_date,
        'rows': rows,
        'has_data': has_data,
        'section_label': section_label,
    }
    return render(request, 'voyage/indices_dashboard.html', context)


def indices_chart_data(request, vessel):
    from collections import defaultdict
    vessel_key = vessel.lower()
    if vessel_key not in VESSEL_INDEX_GROUPS:
        return JsonResponse({'error': 'Unknown vessel'}, status=404)

    today = timezone.localdate()
    default_start = today - timedelta(days=365)

    try:
        start_date = datetime.strptime(request.GET.get('start', default_start.isoformat()), '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start

    try:
        end_date = datetime.strptime(request.GET.get('end', today.isoformat()), '%Y-%m-%d').date()
    except ValueError:
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    indices_param = request.GET.get('indices', '')
    requested = [i.strip() for i in indices_param.split(',') if i.strip()] if indices_param else []

    all_vessel_indices = _indices_for_vessel(vessel_key)
    allowed = set(all_vessel_indices)
    selected = [i for i in requested if i in allowed] if requested else list(all_vessel_indices)

    if not selected:
        return JsonResponse({'traces': [], 'indices': []})

    qs = (
        DailyIndexValue.objects
        .filter(index__name__in=selected, index__is_active=True, date__gte=start_date, date__lte=end_date)
        .select_related('index')
        .order_by('index__name', 'date')
        .values_list('index__name', 'date', 'value')
    )

    buckets = defaultdict(lambda: {'x': [], 'y': []})
    for index_name, date_val, value in qs:
        buckets[index_name]['x'].append(date_val.isoformat())
        buckets[index_name]['y'].append(float(value) if value is not None else None)

    traces = [
        {'name': name, 'x': buckets[name]['x'], 'y': buckets[name]['y']}
        for name in selected if name in buckets
    ]
    return JsonResponse({'traces': traces, 'indices': selected})


def indices_custom(request):
    all_indices = _all_known_indices()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_preset':
            preset_name = (request.POST.get('preset_name') or '').strip()
            selected_raw = request.POST.getlist('indices')
            selected_indices = []
            seen = set()
            for index_name in selected_raw:
                cleaned = index_name.strip()
                if not cleaned or len(cleaned) > 80 or cleaned in seen:
                    continue
                seen.add(cleaned)
                selected_indices.append(cleaned)

            if preset_name and selected_indices:
                CustomIndexPreset.objects.update_or_create(
                    name=preset_name,
                    defaults={'indices': selected_indices},
                )

        if action == 'delete_preset':
            preset_id = request.POST.get('preset_id')
            try:
                CustomIndexPreset.objects.filter(id=int(preset_id)).delete()
            except (TypeError, ValueError):
                pass

        return redirect('voyage:indices_custom')

    presets = list(CustomIndexPreset.objects.all())
    selected_preset_id_raw = request.GET.get('preset')
    selected_preset = None
    if selected_preset_id_raw:
        try:
            selected_preset = CustomIndexPreset.objects.filter(id=int(selected_preset_id_raw)).first()
        except ValueError:
            selected_preset = None

    today = timezone.localdate()
    default_start = today - timedelta(days=30)
    start_date_raw = request.GET.get('start_date', default_start.isoformat())
    end_date_raw = request.GET.get('end_date', today.isoformat())

    try:
        start_date = timezone.datetime.strptime(start_date_raw, '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start

    try:
        end_date = timezone.datetime.strptime(end_date_raw, '%Y-%m-%d').date()
    except ValueError:
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    if (end_date - start_date).days > 365:
        start_date = end_date - timedelta(days=365)

    selected_indices_raw = request.GET.getlist('indices')
    selected_indices = []
    seen = set()
    for index_name in selected_indices_raw:
        cleaned = index_name.strip()
        if not cleaned or len(cleaned) > 80 or cleaned in seen:
            continue
        seen.add(cleaned)
        selected_indices.append(cleaned)

    if not selected_indices and selected_preset:
        selected_indices = selected_preset.indices
    if not selected_indices:
        selected_indices = all_indices[:8]

    rows, has_data = _build_rows(start_date, end_date, selected_indices)

    context = {
        'all_indices': all_indices,
        'selected_indices': selected_indices,
        'start_date': start_date,
        'end_date': end_date,
        'rows': rows,
        'has_data': has_data,
        'presets': presets,
        'selected_preset_id': selected_preset.id if selected_preset else None,
        'selected_preset_name': selected_preset.name if selected_preset else '',
    }
    return render(request, 'voyage/indices_custom.html', context)


def _resolve_intake(voyage):
    if voyage.intake_mode == 'manual':
        return max(voyage.intake_manual, 0)

    vessel = voyage.vessel
    if not vessel:
        return 0

    draft_limit = voyage.draft_limit or vessel.draft
    draft_ratio = min(draft_limit / vessel.draft, 1) if vessel.draft else 1
    draft_adjusted_dwt = vessel.dwt * draft_ratio

    if voyage.stowage_factor and vessel.grain_capacity:
        capacity_limited = vessel.grain_capacity / voyage.stowage_factor
        return max(min(draft_adjusted_dwt, capacity_limited), 0)

    return max(draft_adjusted_dwt, 0)


def _select_speed_profile(voyage):
    if voyage.speed_profile:
        return voyage.speed_profile

    default_profile = voyage.vessel.speed_profiles.filter(is_default=True).first()
    if default_profile:
        return default_profile

    return voyage.vessel.speed_profiles.order_by('name').first()


def _total_consumption_by_mode(voyage):
    fuel_profile = voyage.fuel_profile or voyage.vessel.fuel_profiles.filter(is_default=True).first()
    if not fuel_profile:
        fuel_profile = voyage.vessel.fuel_profiles.order_by('name').first()

    lines = list(VesselFuelConsumption.objects.filter(fuel_profile=fuel_profile)) if fuel_profile else []
    sea_total = sum(line.sea_consumption for line in lines)
    port_total = sum(line.port_consumption for line in lines)
    return sea_total, port_total


def _blended_fuel_price(voyage, date_value):
    splits = list(voyage.fuel_splits.select_related('fuel_index').all())
    if not splits:
        return None

    lookup = {
        value.index_id: value.value
        for value in DailyIndexValue.objects.filter(
            index_id__in=[line.fuel_index_id for line in splits],
            date=date_value,
        )
    }

    total_weight = 0.0
    total_price = 0.0
    for line in splits:
        index_price = lookup.get(line.fuel_index_id)
        if index_price is None or line.weight_pct <= 0:
            continue
        total_weight += line.weight_pct
        total_price += index_price * line.weight_pct

    if total_weight <= 0:
        return None

    return total_price / total_weight


def freight_matrix(request):
    today = timezone.localdate()
    default_start = today - timedelta(days=30)

    start_date_raw = request.GET.get('start_date', default_start.isoformat())
    end_date_raw = request.GET.get('end_date', today.isoformat())

    try:
        start_date = timezone.datetime.strptime(start_date_raw, '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start

    try:
        end_date = timezone.datetime.strptime(end_date_raw, '%Y-%m-%d').date()
    except ValueError:
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    date_rows = [
        end_date - timedelta(days=offset)
        for offset in range((end_date - start_date).days + 1)
        if (end_date - timedelta(days=offset)).weekday() < 5
    ]

    voyages = list(
        FreightVoyage.objects.filter(is_active=True)
        .select_related('vessel', 'speed_profile', 'fuel_profile', 'daily_hire_index')
        .prefetch_related('fuel_splits__fuel_index', 'vessel__speed_profiles', 'vessel__fuel_profiles')
        .order_by('name')
    )

    warnings = []
    
    if not voyages:
        warnings.append('No active voyages configured. Add voyages in Django admin.')
        logger.warning('Freight matrix: No active voyages found')
    
    hire_index_ids = [v.daily_hire_index_id for v in voyages if v.daily_hire_index_id]
    voyages_without_hire_index = [v.name for v in voyages if not v.daily_hire_index_id]
    if voyages_without_hire_index:
        warnings.append(f'Voyages without daily hire index assignment: {", ".join(voyages_without_hire_index)}')
        logger.warning(f'Freight matrix: Voyages without hire index: {voyages_without_hire_index}')
    
    hire_values = DailyIndexValue.objects.filter(index_id__in=hire_index_ids, date__gte=start_date, date__lte=end_date)
    hire_map = {(v.index_id, v.date): v.value for v in hire_values}
    
    if hire_index_ids and not hire_map:
        warnings.append(f'No daily hire index values found for date range {start_date} to {end_date}. Check uploaded index data.')
        logger.warning(f'Freight matrix: No hire values found for indices {hire_index_ids} in date range')
    # Prepare a time-series cache for relevant index ids so we can fall back
    # to the most recent available value on or before a given date.
    needed_index_ids = set(hire_index_ids)
    for v in voyages:
        for line in v.fuel_splits.all():
            if line.fuel_index_id:
                needed_index_ids.add(line.fuel_index_id)

    index_dates = {}
    index_values = {}
    if needed_index_ids:
        qs = DailyIndexValue.objects.filter(index_id__in=needed_index_ids, date__lte=end_date).order_by('date')
        for val in qs:
            idx = val.index_id
            index_dates.setdefault(idx, []).append(val.date)
            index_values.setdefault(idx, []).append(val.value)
    matrix_rows = []
    # helper to find latest value for index_id on or before date_value
    def _latest_value(idx_id, date_value):
        dates = index_dates.get(idx_id)
        if not dates:
            return None
        # find rightmost date <= date_value
        pos = bisect_right(dates, date_value) - 1
        if pos < 0:
            return None
        return index_values[idx_id][pos]

    for row_date in date_rows:
        rates = {}
        for voyage in voyages:
            if not voyage.daily_hire_index_id:
                rates[voyage.name] = None
                continue

            target_tce = hire_map.get((voyage.daily_hire_index_id, row_date))
            if target_tce is None and voyage.daily_hire_index_id:
                target_tce = _latest_value(voyage.daily_hire_index_id, row_date)

            blended_fuel_price = _blended_fuel_price(voyage, row_date)
            # if blended price is missing for the exact date, try to compute
            # using the latest available fuel index values on or before row_date
            if blended_fuel_price is None:
                splits = list(voyage.fuel_splits.all())
                if splits:
                    total_weight = 0.0
                    total_price = 0.0
                    for line in splits:
                        idx_price = _latest_value(line.fuel_index_id, row_date)
                        if idx_price is None or line.weight_pct <= 0:
                            continue
                        total_weight += line.weight_pct
                        total_price += idx_price * line.weight_pct
                    if total_weight > 0:
                        blended_fuel_price = total_price / total_weight
            speed = _select_speed_profile(voyage)

            # track whether we used fallback (earlier-date) values
            hire_used_fallback = False
            fuel_used_fallback = False

            if (
                target_tce is None
                or blended_fuel_price is None
                or not speed
                or speed.ballast_speed <= 0
                or speed.laden_speed <= 0
            ):
                # try fallbacks where possible
                if target_tce is None and voyage.daily_hire_index_id:
                    target_tce = _latest_value(voyage.daily_hire_index_id, row_date)
                    if target_tce is not None:
                        hire_used_fallback = True

                if blended_fuel_price is None:
                    splits = list(voyage.fuel_splits.all())
                    if splits:
                        total_weight = 0.0
                        total_price = 0.0
                        for line in splits:
                            idx_price = _latest_value(line.fuel_index_id, row_date)
                            if idx_price is None or line.weight_pct <= 0:
                                continue
                            total_weight += line.weight_pct
                            total_price += idx_price * line.weight_pct
                        if total_weight > 0:
                            blended_fuel_price = total_price / total_weight
                            fuel_used_fallback = True

            # if still missing critical inputs, skip
            if (
                target_tce is None
                or blended_fuel_price is None
                or not speed
                or speed.ballast_speed <= 0
                or speed.laden_speed <= 0
            ):
                rates[voyage.name] = None
                continue

            intake = _resolve_intake(voyage)
            if intake <= 0 or voyage.load_rate <= 0 or voyage.discharge_rate <= 0:
                rates[voyage.name] = None
                continue

            sea_consumption, port_consumption = _total_consumption_by_mode(voyage)
            if sea_consumption <= 0:
                rates[voyage.name] = None
                continue

            ballast_margin = voyage.sea_margin_ballast_pct / 100.0
            laden_margin = (
                voyage.sea_margin_ballast_pct / 100.0
                if voyage.apply_same_sea_margin
                else voyage.sea_margin_laden_pct / 100.0
            )

            ballast_days = ((voyage.ballast_distance / speed.ballast_speed) / 24.0) * (1 + ballast_margin)
            laden_days = ((voyage.laden_distance / speed.laden_speed) / 24.0) * (1 + laden_margin)
            load_days = (intake / voyage.load_rate) + (voyage.turntime_load_hours / 24.0)
            discharge_days = (intake / voyage.discharge_rate) + (voyage.turntime_discharge_hours / 24.0)
            voyage_days = ballast_days + laden_days + load_days + discharge_days

            if voyage_days <= 0:
                rates[voyage.name] = None
                continue

            fuel_mt = ((ballast_days + laden_days) * sea_consumption) + ((load_days + discharge_days) * port_consumption)
            total_port_expenses = voyage.port_exp_load_port + voyage.port_exp_discharge_port + voyage.misc_expenses
            commission = (voyage.address_commission_pct + voyage.brokerage_commission_pct) / 100.0

            denominator = intake * (1 - commission)
            if denominator <= 0:
                rates[voyage.name] = None
                continue

            freight_rate = ((target_tce * voyage_days) + total_port_expenses + (fuel_mt * blended_fuel_price)) / denominator
            rates[voyage.name] = {
                'value': round(freight_rate, 2),
                'asterisk': bool(hire_used_fallback or fuel_used_fallback),
            }

        matrix_rows.append({'date': row_date, 'rates': rates})

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'voyages': voyages,
        'matrix_rows': matrix_rows,
        'warnings': warnings,
    }
    return render(request, 'voyage/freight_matrix.html', context)


def tce_calculator(request):
    """
    Main TCE Calculator viewer.
    Handles route selection and TCE/Freight calculations.
    """
    try:
        calculated_tce = None
        calculated_freight = None
        db_not_ready = False
        db_error_message = ""
        routes = []
        route_choices = []

        try:
            routes = list(RouteParameters.objects.all())
            route_choices = [(str(route.id), route.route) for route in routes]
        except (ProgrammingError, OperationalError) as exc:
            routes = []
            route_choices = []
            db_not_ready = True
            db_error_message = str(exc)

        form_initial = {}

        if request.method == 'POST' and not db_not_ready:
            post_data = request.POST.copy()
            route_id = post_data.get('route')

            if route_id:
                try:
                    route = RouteParameters.objects.get(id=route_id)
                    route_values = {
                        'route': str(route.id),
                        'ballast_distance': route.ballast_distance,
                        'laden_distance': route.laden_distance,
                        'intake': route.intake,
                        'load_rate': route.load_rate,
                        'discharge_rate': route.discharge_rate,
                        'turntime_hours': route.turntime_hours,
                        'port_exp_load_port': route.port_exp_load_port,
                        'port_exp_discharge_port': route.port_exp_discharge_port,
                        'freight_commission_pct': route.freight_commission_pct,
                        'sea_margin_pct': route.sea_margin_pct,
                        'ballast_speed': route.ballast_speed,
                        'laden_speed': route.laden_speed,
                        'ballast_consumption': route.ballast_consumption,
                        'laden_consumption': route.laden_consumption,
                        'port_consumption': route.port_consumption,
                    }
                    form_initial.update(route_values)

                    if 'calc_tce' not in post_data and 'calc_freight' not in post_data:
                        for key, value in route_values.items():
                            post_data[key] = value
                except (RouteParameters.DoesNotExist, DatabaseError):
                    pass

            form = TCECalculatorForm(post_data, route_choices=route_choices)

            if form.is_valid():
                # Calculate common values used by both calculators.
                data = form.cleaned_data
                common_data = calculate_fuel_and_days(
                    ballast_distance=data['ballast_distance'],
                    laden_distance=data['laden_distance'],
                    intake=data['intake'],
                    load_rate=data['load_rate'],
                    discharge_rate=data['discharge_rate'],
                    turntime_hours=data['turntime_hours'],
                    port_exp_load_port=data['port_exp_load_port'],
                    port_exp_discharge_port=data['port_exp_discharge_port'],
                    freight_commission_pct=data['freight_commission_pct'],
                    sea_margin_pct=data['sea_margin_pct'],
                    ballast_speed=data['ballast_speed'],
                    laden_speed=data['laden_speed'],
                    ballast_consumption=data['ballast_consumption'],
                    laden_consumption=data['laden_consumption'],
                    port_consumption=data['port_consumption'],
                )

                if 'calc_tce' in request.POST:
                    calculated_tce = calculate_tce(
                        freight_rate=data['freight_rate'],
                        fuel_price=data['fuel_price'],
                        intake=data['intake'],
                        common_data=common_data
                    )
                    form_initial['tce_field'] = calculated_tce
                elif 'calc_freight' in request.POST:
                    calculated_freight = calculate_freight_from_tce(
                        target_tce=data['tce_field'],
                        fuel_price=data['fuel_price'],
                        intake=data['intake'],
                        common_data=common_data
                    )
                    form_initial['freight_rate'] = calculated_freight
        else:
            if routes:
                first_route = routes[0]
                form_initial = {
                    'route': str(first_route.id),
                    'ballast_distance': first_route.ballast_distance,
                    'laden_distance': first_route.laden_distance,
                    'intake': first_route.intake,
                    'load_rate': first_route.load_rate,
                    'discharge_rate': first_route.discharge_rate,
                    'turntime_hours': first_route.turntime_hours,
                    'port_exp_load_port': first_route.port_exp_load_port,
                    'port_exp_discharge_port': first_route.port_exp_discharge_port,
                    'freight_commission_pct': first_route.freight_commission_pct,
                    'sea_margin_pct': first_route.sea_margin_pct,
                    'ballast_speed': first_route.ballast_speed,
                    'laden_speed': first_route.laden_speed,
                    'ballast_consumption': first_route.ballast_consumption,
                    'laden_consumption': first_route.laden_consumption,
                    'port_consumption': first_route.port_consumption,
                }
            form = TCECalculatorForm(route_choices=route_choices)

        for key, value in form_initial.items():
            if key in form.fields:
                form.fields[key].initial = value

        context = {
            'form': form,
            'routes': routes,
            'calculated_tce': calculated_tce,
            'calculated_freight': calculated_freight,
            'db_not_ready': db_not_ready,
            'db_error_message': db_error_message,
        }
        return render(request, 'voyage/tce_calculator.html', context)
    except Exception as exc:
        logger.exception("Unhandled error in tce_calculator")
        form = TCECalculatorForm(route_choices=[])
        context = {
            'form': form,
            'routes': [],
            'calculated_tce': None,
            'calculated_freight': None,
            'db_not_ready': True,
            'db_error_message': str(exc),
        }
        return render(request, 'voyage/tce_calculator.html', context, status=200)


@staff_member_required(login_url='/admin/login/')
def upload_pdf_indices(request):
    vessel_size_choices = [
        ('capesize', 'Capesize'),
        ('panamax', 'Panamax'),
        ('supramax', 'Supramax'),
        ('handysize', 'Handysize'),
        ('bunker', 'Bunker'),
    ]

    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        vessel_size = request.POST.get('vessel_size', 'panamax')
        pages = request.POST.get('pages', 'all').strip() or 'all'

        if not upload_file:
            messages.error(request, 'Please select a PDF file to upload.')
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        try:
            extracted_tables = _extract_pdf_index_tables(
                upload_file.read(), vessel_size, pages
            )
        except Exception as exc:
            messages.error(request, f'Unable to extract PDF indices: {exc}')
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        if not extracted_tables:
            messages.warning(
                request,
                'No index tables were found in the uploaded PDF. Please check the file and try again.',
            )
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        session_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_pdf_upload')
        os.makedirs(temp_dir, exist_ok=True)
        temp_file = os.path.join(temp_dir, f'{session_id}.json')

        session_data = {
            'file_path': upload_file.name,
            'vessel_size': vessel_size,
            'pages': pages,
            'extracted_tables': extracted_tables,
            'extraction_time': datetime.now().isoformat(),
            'session_id': session_id,
        }

        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(session_data, f, indent=2, default=str)

        return redirect('voyage:verify_pdf_indices', session_id=session_id)

    return render(
        request,
        'voyage/upload_pdf_indices.html',
        {
            'vessel_size_choices': vessel_size_choices,
            'selected_vessel_size': 'panamax',
            'pages': 'all',
        },
    )


def _parse_pdf_pages(pages_str):
    pages_str = str(pages_str or 'all').strip().lower()
    if pages_str in ('all', ''):
        return 'all'

    parsed_pages = []
    for part in pages_str.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            start, end = part.split('-', 1)
            parsed_pages.extend(range(int(start.strip()), int(end.strip()) + 1))
        else:
            parsed_pages.append(int(part))

    return sorted(set(parsed_pages))


def _extract_pdf_index_tables(file_bytes, vessel_size, pages):
    page_numbers = _parse_pdf_pages(pages)

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        if page_numbers == 'all':
            pages_to_process = list(pdf.pages)
        else:
            pages_to_process = [
                pdf.pages[i - 1]
                for i in page_numbers
                if 1 <= i <= len(pdf.pages)
            ]

        if not pages_to_process:
            return []

        existing_indices = {
            idx.name.strip().lower(): idx
            for idx in AvailableIndex.objects.filter(
                vessel_size=vessel_size,
                is_active=True,
            )
        }

        # Try to extract date from PDF text (usually in header)
        document_date = None
        for page in pages_to_process:
            page_text = page.extract_text() or ''
            for line in page_text.split('\n')[:20]:  # Check first 20 lines for date
                parsed = _parse_date(line.strip())
                if parsed:
                    document_date = parsed
                    break
            if document_date:
                break

        extracted_tables = []
        for page_index, page in enumerate(pages_to_process, start=1):
            tables = page.extract_tables()
            if not tables:
                continue

            for table_index, table in enumerate(tables):
                if not table or len(table) < 2:
                    continue

                cleaned_table = [
                    [str(cell).strip() if cell is not None else '' for cell in row]
                    for row in table
                    if row is not None
                ]

                header_row_idx, headers = _find_header_row(cleaned_table)
                if header_row_idx is None or not headers:
                    continue

                # Find date column - could be "Date", "Period", or something like "Today"
                date_col_idx = None
                date_col_is_explicit = False
                for idx, header in enumerate(headers):
                    if header and any(
                        keyword in header.lower()
                        for keyword in ['date', 'period', 'month', 'day']
                    ):
                        date_col_idx = idx
                        date_col_is_explicit = True
                        break

                # If no explicit date column, use the first column as the index name
                if date_col_idx is None:
                    date_col_idx = 0

                index_columns = []
                for idx, header in enumerate(headers):
                    if idx == date_col_idx and date_col_is_explicit:
                        continue
                    if not header or header.lower() in ['change', 'change %', 'change%', '%', 'mtd', 'ytd', 'today']:
                        continue
                    normalized = header.strip()
                    if len(normalized) < 2:
                        continue
                    index_columns.append(
                        {
                            'name': normalized,
                            'column_index': idx,
                            'existing_index': normalized.strip().lower() in existing_indices,
                        }
                    )

                if not index_columns:
                    continue

                processed_rows = []
                for row in cleaned_table[header_row_idx + 1 :]:
                    if len(row) < 1:
                        continue

                    # Get date from row or use document date
                    parsed_date = document_date
                    if date_col_is_explicit and len(row) > date_col_idx:
                        date_str = row[date_col_idx].strip()
                        if date_str:
                            parsed = _parse_date(date_str)
                            if parsed:
                                parsed_date = parsed

                    if not parsed_date:
                        continue

                    # Index name is usually in the first column
                    index_name = row[0].strip() if len(row) > 0 else ''
                    if not index_name or len(index_name) < 2:
                        continue

                    indices = {}

                    # Try to get the "Today" value (usually second column if first is name)
                    # or first numeric column if the date is explicit
                    for idx_col in index_columns:
                        col_idx = idx_col['column_index']
                        if col_idx >= len(row):
                            continue
                        raw_value = row[col_idx]
                        if raw_value is None:
                            continue
                        value_text = str(raw_value).strip()
                        if not value_text or value_text.lower() in ['change', 'change %', 'change%', '%', 'mtd', 'ytd']:
                            continue

                        cleaned_value = (
                            value_text.replace(',', '')
                            .replace('$', '')
                            .replace('£', '')
                            .replace('€', '')
                            .strip()
                        )

                        try:
                            value = float(cleaned_value)
                        except ValueError:
                            continue

                        # Use the row's first column as the index name if header was generic
                        final_index_name = idx_col['name']
                        if idx_col['column_index'] == 0:
                            final_index_name = index_name

                        indices[final_index_name] = {
                            'value': value,
                            'original_value': value_text,
                            'existing': final_index_name.strip().lower() in existing_indices,
                        }

                    if indices:
                        processed_rows.append(
                            {
                                'date': parsed_date.isoformat(),
                                'original_date': parsed_date.strftime('%d %b %Y'),
                                'indices': indices,
                            }
                        )

                if not processed_rows:
                    continue

                extracted_tables.append(
                    {
                        'page': page_index,
                        'table_index': table_index,
                        'data': {
                            'headers': headers,
                            'date_column': date_col_idx,
                            'index_columns': index_columns,
                            'data': processed_rows,
                            'total_rows': len(processed_rows),
                        },
                        'raw_table': cleaned_table[:10],
                    }
                )

        return extracted_tables


def _find_header_row(cleaned_table):
    for row_index, row in enumerate(cleaned_table):
        if any(
            cell
            and any(keyword in cell.lower() for keyword in ['date', 'period', 'month', 'day'])
            for cell in row
        ):
            return row_index, row
    return (0, cleaned_table[0]) if cleaned_table else (None, [])


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None

        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%m-%d-%Y',
            '%b %d, %Y',
            '%B %d, %Y',
            '%d %b %Y',
            '%d %B %Y',
            '%Y/%m/%d',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue

        normalized = cleaned.replace('/', '-').replace('.', '-').lower()
        month_names = {
            'jan': '01',
            'feb': '02',
            'mar': '03',
            'apr': '04',
            'may': '05',
            'jun': '06',
            'jul': '07',
            'aug': '08',
            'sep': '09',
            'oct': '10',
            'nov': '11',
            'dec': '12',
        }
        for month_name, month_num in month_names.items():
            if month_name in normalized:
                normalized = normalized.replace(month_name, month_num)

        for fmt in formats:
            try:
                return datetime.strptime(normalized, fmt).date()
            except ValueError:
                continue

    return None


def verify_pdf_indices(request, session_id):
    """
    View for verifying and validating extracted PDF indices before adding to database.
    """
    temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_pdf_upload')
    temp_file = os.path.join(temp_dir, f'{session_id}.json')

    if not os.path.exists(temp_file):
        messages.error(request, 'Verification session not found or expired.')
        return redirect('voyage:upload_pdf_indices')

    try:
        with open(temp_file, 'r') as f:
            session_data = json.load(f)
    except Exception as e:
        messages.error(request, f'Error loading verification data: {e}')
        return redirect('voyage:upload_pdf_indices')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_selected':
            return _save_selected_indices(request, session_data, temp_file)
        elif action == 'discard':
            # Remove temp file
            try:
                os.remove(temp_file)
            except:
                pass
            messages.info(request, 'PDF upload cancelled.')
            return redirect('voyage:upload_pdf_indices')

    # Prepare data for template
    extracted_tables = session_data.get('extracted_tables', [])
    vessel_size = session_data.get('vessel_size', 'panamax')

    # Get existing indices for validation
    existing_indices = {
        idx.name.lower(): idx.name
        for idx in AvailableIndex.objects.filter(vessel_size=vessel_size, is_active=True)
    }

    context = {
        'session_id': session_id,
        'file_path': session_data.get('file_path', ''),
        'vessel_size': vessel_size,
        'extraction_time': session_data.get('extraction_time', ''),
        'extracted_tables': extracted_tables,
        'existing_indices': existing_indices,
        'total_tables': len(extracted_tables),
    }

    return render(request, 'voyage/verify_pdf_indices.html', context)


def _save_selected_indices(request, session_data, temp_file):
    """
    Save selected indices from the verification form.
    """
    selected_tables = request.POST.getlist('selected_tables')
    vessel_size = session_data.get('vessel_size', 'panamax')
    extracted_tables = session_data.get('extracted_tables', [])

    if not selected_tables:
        messages.warning(request, 'No tables selected for import.')
        return redirect(request.META.get('HTTP_REFERER', reverse('voyage:upload_pdf_indices')))

    saved_count = 0
    error_count = 0

    try:
        with transaction.atomic():
            for table_idx_str in selected_tables:
                try:
                    table_idx = int(table_idx_str)
                    if table_idx >= len(extracted_tables):
                        continue

                    table_data = extracted_tables[table_idx]

                    # Create missing indices
                    index_columns = table_data['data']['index_columns']
                    for idx_col in index_columns:
                        if not idx_col.get('existing_index'):
                            # Create new index
                            next_order = AvailableIndex.objects.filter(vessel_size=vessel_size).aggregate(
                                Max('order')
                            )['order__max'] or 0
                            AvailableIndex.objects.create(
                                name=idx_col['name'],
                                vessel_size=vessel_size,
                                order=next_order + 1,
                                is_active=True
                            )

                    # Save daily values
                    from .models import DailyIndexValue
                    values_to_create = []

                    for row_data in table_data['data']['data']:
                        row_date = datetime.fromisoformat(row_data['date']).date()

                        for idx_name, idx_data in row_data['indices'].items():
                            try:
                                index_obj = AvailableIndex.objects.get(
                                    name__iexact=idx_name,
                                    vessel_size=vessel_size
                                )
                                values_to_create.append(
                                    DailyIndexValue(
                                        index=index_obj,
                                        date=row_date,
                                        value=idx_data['value']
                                    )
                                )
                            except AvailableIndex.DoesNotExist:
                                continue

                    # Bulk create (ignore conflicts for existing data)
                    DailyIndexValue.objects.bulk_create(values_to_create, ignore_conflicts=True)
                    saved_count += len(values_to_create)

                except Exception as e:
                    error_count += 1
                    continue

    except Exception as e:
        messages.error(request, f'Error saving indices: {e}')
        return redirect(request.META.get('HTTP_REFERER', reverse('voyage:upload_pdf_indices')))

    # Clean up temp file
    try:
        os.remove(temp_file)
    except:
        pass

    if saved_count > 0:
        messages.success(request, f'Successfully imported {saved_count} index values.')
    if error_count > 0:
        messages.warning(request, f'{error_count} tables had errors during import.')

    return redirect('voyage:upload_pdf_indices')


def _rate_code_vessel_size(code):
    c = (code or '').upper().strip()
    if c.startswith('HS'):
        return 'handysize'
    if c.startswith('C') or c.startswith('BCI'):
        return 'capesize'
    if c.startswith('P') or c.startswith('BPI'):
        return 'panamax'
    if c.startswith('S') or c.startswith('BSI'):
        return 'supramax'
    return 'panamax'


def _parse_excel_spot_rows(upload_file):
    """Parse the Baltic tab of an uploaded .xlsx and return (spot_rows, error_msg).

    spot_rows is a list of dicts: {code, description, date (date obj), value (float)}.
    Returns (None, error_msg) on failure.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
    except Exception as exc:
        return None, f'Could not open file: {exc}'

    if 'Baltic' not in wb.sheetnames:
        return None, "No 'Baltic' sheet found in this workbook."

    ws = wb['Baltic']
    spot_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        rate_code, desc, period, rate_date, value = row[0], row[1], row[2], row[3], row[4]
        if period != 'Spot':
            continue
        if not rate_code or value is None:
            continue
        if isinstance(rate_date, datetime):
            rate_date = rate_date.date()
        elif isinstance(rate_date, str):
            rate_date = _parse_date(rate_date)
        if not rate_date:
            continue
        try:
            value = float(value)
        except (TypeError, ValueError):
            continue
        spot_rows.append({
            'code': str(rate_code).strip(),
            'description': str(desc).strip() if desc else '',
            'date': rate_date.isoformat(),
            'value': value,
        })
    if not spot_rows:
        return None, 'No Spot rows found in the Baltic sheet.'
    return spot_rows, None


@staff_member_required(login_url='/admin/login/')
def upload_excel_indices(request):
    """Upload the daily Baltic Exchange End-of-Day .xlsx — parse and redirect to mapping review."""
    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        if not upload_file:
            messages.error(request, 'Please select an .xlsx file.')
            return render(request, 'voyage/upload_excel_indices.html', {})

        spot_rows, err = _parse_excel_spot_rows(upload_file)
        if err:
            messages.error(request, err)
            return render(request, 'voyage/upload_excel_indices.html', {})

        session_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
        os.makedirs(temp_dir, exist_ok=True)
        temp_file = os.path.join(temp_dir, f'{session_id}.json')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump({'filename': upload_file.name, 'rows': spot_rows}, f)

        return redirect('voyage:review_excel_mappings', session_id=session_id)

    return render(request, 'voyage/upload_excel_indices.html', {})


@staff_member_required(login_url='/admin/login/')
def upload_batch_indices(request):
    """Upload a historical columnar indices .xlsx (Date column + one column per index).

    Only reads the header row during the upload request — the full file is saved to
    disk and processed during the confirm step to avoid request timeouts on large files.
    """
    VESSEL_CHOICES = [
        ('panamax', 'Panamax'),
        ('capesize', 'Capesize'),
        ('supramax', 'Supramax'),
        ('handysize', 'Handysize'),
        ('bunker', 'Bunker'),
    ]

    if request.method == 'POST':
        import openpyxl
        upload_file = request.FILES.get('upload_file')
        vessel_size = request.POST.get('vessel_size', 'panamax')
        if vessel_size not in dict(VESSEL_CHOICES):
            vessel_size = 'panamax'

        if not upload_file:
            messages.error(request, 'Please select an .xlsx file.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
        os.makedirs(temp_dir, exist_ok=True)
        session_id = str(uuid.uuid4())

        # Save the raw file to disk — avoid loading all rows in the web request
        xlsx_path = os.path.join(temp_dir, f'{session_id}.xlsx')
        with open(xlsx_path, 'wb') as fh:
            for chunk in upload_file.chunks():
                fh.write(chunk)

        # Read only the header row to get column names
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as exc:
            os.remove(xlsx_path)
            messages.error(request, f'Could not open file: {exc}')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()

        if not header or str(header[0]).strip().lower() != 'date':
            os.remove(xlsx_path)
            messages.error(request, 'Expected first column to be "Date". Check the file format.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        columns = [str(c).strip() for c in header[1:] if c is not None and str(c).strip()]
        if not columns:
            os.remove(xlsx_path)
            messages.error(request, 'No index columns found in header row.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        # Store session metadata — full parsing happens at confirm time
        meta_file = os.path.join(temp_dir, f'{session_id}.json')
        with open(meta_file, 'w', encoding='utf-8') as f:
            json.dump({
                'type': 'batch',
                'filename': upload_file.name,
                'vessel_size': vessel_size,
                'xlsx_path': xlsx_path,
                'columns': columns,
            }, f)

        return redirect('voyage:review_excel_mappings', session_id=session_id)

    return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})


@staff_member_required(login_url='/admin/login/')
def review_excel_mappings(request, session_id):
    """Show / confirm per-code mapping between Excel RateCodes and AvailableIndex entries.

    Handles two session types:
      'batch' — large historical files saved as .xlsx on disk; parsed at confirm time
               using streaming + batched bulk_create to avoid timeouts.
      'eod'   — small daily EOD files whose rows were pre-parsed into the session JSON.
    """
    import openpyxl
    from .models import IndexCodeMapping

    temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
    temp_file = os.path.join(temp_dir, f'{session_id}.json')

    if not os.path.exists(temp_file):
        messages.error(request, 'Upload session not found or expired. Please upload again.')
        return redirect('voyage:upload_excel_indices')

    with open(temp_file, 'r', encoding='utf-8') as f:
        session_data = json.load(f)

    filename = session_data.get('filename', '')
    session_type = session_data.get('type', 'eod')       # 'batch' or 'eod'
    session_vessel_size = session_data.get('vessel_size') # may be None for EOD

    # Build unique-codes dict for the review table.
    # Batch sessions store only column names (no row data yet).
    # EOD sessions store pre-parsed rows.
    seen = {}  # code -> representative dict {code, description, date, value}
    if session_type == 'batch':
        for col in session_data.get('columns', []):
            if col not in seen:
                seen[col] = {'code': col, 'description': col, 'date': '—', 'value': ''}
    else:
        for r in session_data.get('rows', []):
            code = r['code']
            if code not in seen:
                seen[code] = r

    # Common DB lookups
    all_indices = list(AvailableIndex.objects.filter(is_active=True).order_by('vessel_size', 'name'))
    saved_mappings = {m.rate_code: m for m in IndexCodeMapping.objects.select_related('target_index').all()}
    index_by_name = {idx.name.lower(): idx for idx in all_indices}

    # ------------------------------------------------------------------ POST
    if request.method == 'POST':
        # Step 1: persist the user's mapping choices
        with transaction.atomic():
            for code in seen:
                action = request.POST.get(f'action_{code}', 'auto')
                target_id = request.POST.get(f'target_{code}', '')

                if action == 'skip':
                    IndexCodeMapping.objects.update_or_create(
                        rate_code=code,
                        defaults={'skip': True, 'target_index': None},
                    )
                    continue

                target_index = None
                if target_id:
                    try:
                        target_index = AvailableIndex.objects.get(pk=int(target_id))
                    except (AvailableIndex.DoesNotExist, ValueError):
                        pass

                if target_index is None:
                    vs = session_vessel_size or _rate_code_vessel_size(code)
                    target_index, _ = AvailableIndex.objects.get_or_create(
                        name=code,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )

                IndexCodeMapping.objects.update_or_create(
                    rate_code=code,
                    defaults={'skip': False, 'target_index': target_index},
                )

        # Step 2: import values — different strategy per session type
        mappings = {m.rate_code: m for m in IndexCodeMapping.objects.select_related('target_index').all()}
        dates_seen = set()
        new_values = 0

        if session_type == 'batch':
            # Stream the saved .xlsx row-by-row; bulk_create in batches of 500
            xlsx_path = session_data.get('xlsx_path', '')
            if not os.path.exists(xlsx_path):
                messages.error(request, 'Uploaded file no longer found. Please upload again.')
                return redirect('voyage:upload_batch_indices')

            # Resolve index objects for each column up-front
            col_index_map = {}   # col_name -> AvailableIndex (or None if skipped)
            for col in session_data.get('columns', []):
                m = mappings.get(col)
                if m and m.skip:
                    col_index_map[col] = None
                    continue
                idx_obj = m.target_index if (m and m.target_index) else None
                if idx_obj is None:
                    vs = session_vessel_size or _rate_code_vessel_size(col)
                    idx_obj, _ = AvailableIndex.objects.get_or_create(
                        name=col,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )
                col_index_map[col] = idx_obj

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            header = None
            index_cols = []   # [(col_offset_in_header, col_name), ...]
            batch = []
            BATCH_SIZE = 500

            for row in ws.iter_rows(min_row=1, values_only=True):
                if header is None:
                    header = row
                    for i, cell in enumerate(header[1:], start=1):
                        name = str(cell).strip() if cell else ''
                        if name and name in col_index_map:
                            index_cols.append((i, name))
                    continue

                date_val = row[0] if row else None
                if date_val is None:
                    continue
                if isinstance(date_val, datetime):
                    row_date = date_val.date()
                else:
                    row_date = _parse_date(str(date_val))
                if not row_date:
                    continue

                for col_i, col_name in index_cols:
                    idx_obj = col_index_map.get(col_name)
                    if idx_obj is None:
                        continue
                    val = row[col_i] if col_i < len(row) else None
                    if val is None:
                        continue
                    try:
                        val = float(val)
                    except (TypeError, ValueError):
                        continue
                    batch.append(DailyIndexValue(index=idx_obj, date=row_date, value=val))
                    dates_seen.add(row_date)
                    if len(batch) >= BATCH_SIZE:
                        created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                        new_values += len(created)
                        batch = []

            wb.close()
            if batch:
                created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                new_values += len(created)

            # Clean up both temp files
            for path in (temp_file, xlsx_path):
                try:
                    os.remove(path)
                except OSError:
                    pass

        else:
            # EOD: rows already parsed in session JSON — use bulk_create too
            batch = []
            BATCH_SIZE = 500
            for r in session_data.get('rows', []):
                code = r['code']
                m = mappings.get(code)
                if m and m.skip:
                    continue
                idx_obj = m.target_index if (m and m.target_index) else None
                if idx_obj is None:
                    vs = session_vessel_size or _rate_code_vessel_size(code)
                    idx_obj, _ = AvailableIndex.objects.get_or_create(
                        name=code,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )
                row_date = _parse_date(r['date'])
                if not row_date:
                    continue
                batch.append(DailyIndexValue(index=idx_obj, date=row_date, value=r['value']))
                dates_seen.add(row_date)
                if len(batch) >= BATCH_SIZE:
                    created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                    new_values += len(created)
                    batch = []
            if batch:
                created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                new_values += len(created)
            try:
                os.remove(temp_file)
            except OSError:
                pass

        if new_values:
            date_summary = f'{min(dates_seen)} → {max(dates_seen)}' if len(dates_seen) > 3 else ', '.join(str(d) for d in sorted(dates_seen))
            messages.success(request, f'Imported {new_values} index values ({date_summary}).')
        else:
            messages.info(request, 'All values already existed — nothing new imported.')

        redirect_url = 'voyage:upload_batch_indices' if session_type == 'batch' else 'voyage:upload_excel_indices'
        return redirect(redirect_url)

    # ------------------------------------------------------------------ GET
    mapping_rows = []
    for code, rep in seen.items():
        saved = saved_mappings.get(code)
        if saved and not saved.skip and saved.target_index:
            suggested = saved.target_index
            status = 'mapped'
        elif saved and saved.skip:
            suggested = None
            status = 'skip'
        elif code.lower() in index_by_name:
            suggested = index_by_name[code.lower()]
            status = 'exact'
        else:
            suggested = None
            status = 'new'

        mapping_rows.append({
            'code': code,
            'description': rep['description'],
            'date': rep['date'],
            'value': rep['value'],
            'suggested': suggested,
            'status': status,
        })

    mapping_rows.sort(key=lambda r: (0 if r['status'] in ('new', 'skip') else 1, r['code']))

    context = {
        'filename': filename,
        'session_id': session_id,
        'session_type': session_type,
        'mapping_rows': mapping_rows,
        'all_indices': all_indices,
        'total': len(mapping_rows),
        'unmatched': sum(1 for r in mapping_rows if r['status'] == 'new'),
    }
    return render(request, 'voyage/review_excel_mappings.html', context)


def vessel_compare(request):
    from .models import ComparisonVessel, ComparisonVoyage, VesselCompareConfig, VesselVoyageIntake

    def _f(val, default=0.0):
        try:
            return float(val)
        except (ValueError, TypeError):
            return float(default)

    def _run_calc(cfg, voyages_qs, vessels_qs):
        voyages = [
            {'name': voy.name, 'ballast_dist': voy.ballast_dist, 'laden_dist': voy.laden_dist,
             'load_rate': voy.load_rate, 'dis_rate': voy.dis_rate,
             'load_factor': voy.load_factor, 'dis_factor': voy.dis_factor,
             'turntimes_hours': voy.turntimes_hours,
             'port_exp': voy.port_exp, 'various_exp': voy.various_exp}
            for voy in voyages_qs
        ]
        # Build intake map: {vessel_id: {voyage_id: intake}}
        intake_map = {}
        for vi in VesselVoyageIntake.objects.filter(
            vessel__in=vessels_qs, voyage__in=voyages_qs
        ).select_related('vessel', 'voyage'):
            intake_map.setdefault(vi.vessel_id, {})[vi.voyage_id] = vi.intake

        vessels = []
        for v in vessels_qs:
            v_intakes = intake_map.get(v.id, {})
            intakes = [v_intakes.get(voy.id, v.default_intake) for voy in voyages_qs]
            vessels.append({
                'name': v.name, 'intakes': intakes,
                'laden_speed': v.laden_speed, 'ballast_speed': v.ballast_speed,
                'laden_cons': v.laden_cons, 'ballast_cons': v.ballast_cons,
                'port_cons': v.port_cons,
            })

        global_inputs = {'hire': cfg.hire, 'ifo_price': cfg.ifo_price,
                         'mgo_price': cfg.mgo_price, 'weather_factor': cfg.weather_factor}
        r = calculate_vessel_comparison(global_inputs, voyages, vessels)
        r['vessel_summary'] = [
            {'name': name, 'wa': wa,
             'pct': wa * 100 if wa is not None else None,
             'pct_delta': (wa - 1) * 100 if wa is not None else None}
            for name, wa in zip(r['vessels'], r['weighted_avgs'])
        ]
        return r

    cfg = VesselCompareConfig.get()

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'update_config':
            cfg.hire = _f(request.POST.get('hire'), cfg.hire)
            cfg.ifo_price = _f(request.POST.get('ifo_price'), cfg.ifo_price)
            cfg.mgo_price = _f(request.POST.get('mgo_price'), cfg.mgo_price)
            cfg.weather_factor = _f(request.POST.get('weather_factor'), cfg.weather_factor)
            cfg.save()
            return redirect('voyage:vessel_compare')

        elif action == 'add_voyage':
            name = request.POST.get('new_voy_name', '').strip()
            if name:
                next_order = ComparisonVoyage.objects.count()
                voy = ComparisonVoyage.objects.create(
                    name=name, order=next_order,
                    ballast_dist=_f(request.POST.get('new_voy_ballast_dist'), 5000),
                    laden_dist=_f(request.POST.get('new_voy_laden_dist'), 5000),
                    load_rate=_f(request.POST.get('new_voy_load_rate'), 10000),
                    dis_rate=_f(request.POST.get('new_voy_dis_rate'), 10000),
                    load_factor=_f(request.POST.get('new_voy_load_factor'), 1.0),
                    dis_factor=_f(request.POST.get('new_voy_dis_factor'), 1.0),
                    turntimes_hours=_f(request.POST.get('new_voy_turntimes'), 36),
                    port_exp=_f(request.POST.get('new_voy_port_exp'), 100000),
                    various_exp=_f(request.POST.get('new_voy_various_exp'), 10000),
                )
                # Create default intake records for all existing vessels
                default_intake = _f(request.POST.get('new_voy_default_intake'), 79000)
                for vessel in ComparisonVessel.objects.all():
                    VesselVoyageIntake.objects.get_or_create(
                        vessel=vessel, voyage=voy,
                        defaults={'intake': default_intake},
                    )
            return redirect('voyage:vessel_compare')

        elif action == 'edit_voyage':
            vid = request.POST.get('voyage_id')
            try:
                voy = ComparisonVoyage.objects.get(pk=int(vid))
                voy.name = request.POST.get('edit_voy_name', voy.name).strip() or voy.name
                voy.ballast_dist = _f(request.POST.get('edit_voy_ballast_dist'), voy.ballast_dist)
                voy.laden_dist = _f(request.POST.get('edit_voy_laden_dist'), voy.laden_dist)
                voy.load_rate = _f(request.POST.get('edit_voy_load_rate'), voy.load_rate)
                voy.dis_rate = _f(request.POST.get('edit_voy_dis_rate'), voy.dis_rate)
                voy.load_factor = _f(request.POST.get('edit_voy_load_factor'), voy.load_factor)
                voy.dis_factor = _f(request.POST.get('edit_voy_dis_factor'), voy.dis_factor)
                voy.turntimes_hours = _f(request.POST.get('edit_voy_turntimes'), voy.turntimes_hours)
                voy.port_exp = _f(request.POST.get('edit_voy_port_exp'), voy.port_exp)
                voy.various_exp = _f(request.POST.get('edit_voy_various_exp'), voy.various_exp)
                voy.save()
            except (ComparisonVoyage.DoesNotExist, ValueError):
                pass
            return redirect('voyage:vessel_compare')

        elif action == 'delete_voyage':
            vid = request.POST.get('voyage_id')
            try:
                ComparisonVoyage.objects.get(pk=int(vid)).delete()
            except (ComparisonVoyage.DoesNotExist, ValueError):
                pass
            return redirect('voyage:vessel_compare')

        elif action == 'add_vessel':
            name = request.POST.get('new_name', '').strip()
            if name:
                default_intake = _f(request.POST.get('new_default_intake'), 79000)
                next_order = ComparisonVessel.objects.count()
                vessel = ComparisonVessel.objects.create(
                    name=name, order=next_order,
                    default_intake=default_intake,
                    laden_speed=_f(request.POST.get('new_laden_speed'), 12),
                    ballast_speed=_f(request.POST.get('new_ballast_speed'), 12.5),
                    laden_cons=_f(request.POST.get('new_laden_cons'), 22),
                    ballast_cons=_f(request.POST.get('new_ballast_cons'), 23),
                    port_cons=_f(request.POST.get('new_port_cons'), 4.5),
                )
                # Create intake records for each voyage
                for voy in ComparisonVoyage.objects.all():
                    intake_val = _f(request.POST.get(f'new_intake_{voy.id}'), default_intake)
                    VesselVoyageIntake.objects.get_or_create(
                        vessel=vessel, voyage=voy,
                        defaults={'intake': intake_val},
                    )
            return redirect('voyage:vessel_compare')

        elif action == 'edit_vessel':
            vid = request.POST.get('vessel_id')
            try:
                v = ComparisonVessel.objects.get(pk=int(vid))
                v.name = request.POST.get('edit_name', v.name).strip() or v.name
                v.default_intake = _f(request.POST.get('edit_default_intake'), v.default_intake)
                v.laden_speed = _f(request.POST.get('edit_laden_speed'), v.laden_speed)
                v.ballast_speed = _f(request.POST.get('edit_ballast_speed'), v.ballast_speed)
                v.laden_cons = _f(request.POST.get('edit_laden_cons'), v.laden_cons)
                v.ballast_cons = _f(request.POST.get('edit_ballast_cons'), v.ballast_cons)
                v.port_cons = _f(request.POST.get('edit_port_cons'), v.port_cons)
                v.save()
                # Update per-voyage intakes
                for voy in ComparisonVoyage.objects.all():
                    intake_val = _f(request.POST.get(f'edit_intake_{voy.id}'), v.default_intake)
                    VesselVoyageIntake.objects.update_or_create(
                        vessel=v, voyage=voy,
                        defaults={'intake': intake_val},
                    )
            except (ComparisonVessel.DoesNotExist, ValueError):
                pass
            return redirect('voyage:vessel_compare')

        elif action == 'delete_vessel':
            vid = request.POST.get('vessel_id')
            try:
                v = ComparisonVessel.objects.get(pk=int(vid))
                if not v.is_standard:
                    v.delete()
            except (ComparisonVessel.DoesNotExist, ValueError):
                pass
            return redirect('voyage:vessel_compare')

    voyages_qs = ComparisonVoyage.objects.all()
    vessels_qs = ComparisonVessel.objects.all()

    # Build intake lookup for template: {vessel_id: {voyage_id: intake}}
    intake_map = {}
    for vi in VesselVoyageIntake.objects.filter(
        vessel__in=vessels_qs, voyage__in=voyages_qs
    ).select_related('vessel', 'voyage'):
        intake_map.setdefault(vi.vessel_id, {})[vi.voyage_id] = vi.intake

    # Enrich vessels with per-voyage intake list for template
    vessels_with_intakes = []
    for v in vessels_qs:
        v_intakes = intake_map.get(v.id, {})
        vessels_with_intakes.append({
            'obj': v,
            'intakes': {voy.id: v_intakes.get(voy.id, v.default_intake) for voy in voyages_qs},
        })

    results = None
    error = None
    if vessels_qs.exists() and voyages_qs.exists():
        try:
            results = _run_calc(cfg, voyages_qs, vessels_qs)
        except Exception as exc:
            logger.exception("Error in vessel_compare calculation")
            error = str(exc)

    context = {
        'cfg': cfg,
        'voyages': voyages_qs,
        'vessels': vessels_qs,
        'vessels_with_intakes': vessels_with_intakes,
        'results': results,
        'error': error,
    }
    return render(request, 'voyage/vessel_compare.html', context)


def ffa_valuation(request):
    if request.method == 'POST':
        ct = request.content_type or ''
        if 'application/json' in ct:
            body = json.loads(request.body)
        else:
            body = request.POST
        action = body.get('action')

        if action == 'parse':
            result = parse_ffa_text(body.get('raw_text', ''), date.today())
            return JsonResponse({
                'vessel_class': result['vessel_class'],
                'periods': [
                    {
                        'label': p['label'], 'period_type': p['period_type'],
                        'start_date': p['start_date'].isoformat(),
                        'end_date': p['end_date'].isoformat(),
                        'bid': str(p['bid']), 'offer': str(p['offer']),
                    }
                    for p in result['periods']
                ],
            })

        if action == 'save':
            raw_text = body.get('raw_text', '')
            vessel_class = body.get('vessel_class', '')
            periods_data = body.get('periods', [])
            if isinstance(periods_data, str):
                periods_data = json.loads(periods_data)
            curve = FFACurve.objects.create(vessel_class=vessel_class, raw_text=raw_text)
            FFACurvePeriod.objects.bulk_create([
                FFACurvePeriod(
                    curve=curve, label=p['label'], period_type=p['period_type'],
                    start_date=p['start_date'], end_date=p['end_date'],
                    bid=p['bid'], offer=p['offer'],
                )
                for p in periods_data
            ])
            return JsonResponse({'curve_id': curve.id, 'status': 'saved'})

    curve = FFACurve.objects.prefetch_related('periods').first()
    vessels = ComparisonVessel.objects.order_by('name')
    return render(request, 'voyage/ffa_valuation.html', {
        'curve': curve,
        'vessels': vessels,
        'today': date.today(),
    })


def ffa_valuation_calculate(request):
    data = json.loads(request.body)
    delivery_date = date.fromisoformat(data['delivery_date'])
    period_months = int(data.get('period_months', 1))
    vessel_ids = [int(v) for v in data.get('vessel_ids', [])]

    try:
        curve = FFACurve.objects.prefetch_related('periods').get(id=data['curve_id'])
    except FFACurve.DoesNotExist:
        return JsonResponse({'error': 'Curve not found'}, status=404)

    curve_periods = [
        {
            'label': p.label, 'period_type': p.period_type,
            'start_date': p.start_date, 'end_date': p.end_date,
            'bid': p.bid, 'offer': p.offer,
        }
        for p in curve.periods.all()
    ]

    result = resolve_employment_periods(curve_periods, delivery_date, period_months)
    employment_end = result['end_date']

    timeline = [
        {
            'label': p['label'], 'offer': float(p['offer']),
            'start': p['start_date'].isoformat(), 'end': p['end_date'].isoformat(),
            'period_type': p['period_type'],
            'in_window': p['start_date'] < employment_end and p['end_date'] >= delivery_date,
        }
        for p in curve_periods if p['period_type'] != 'combined'
    ]

    vessels_out = []
    if vessel_ids and result['blended_offer'] is not None:
        try:
            from .models import (ComparisonVoyage, VesselCompareConfig,
                                  VesselVoyageIntake)
            cfg = VesselCompareConfig.objects.first()
            voyages_qs = list(ComparisonVoyage.objects.all())
            bki = ComparisonVessel.objects.filter(is_standard=True).first()
            selected = list(ComparisonVessel.objects.filter(
                id__in=vessel_ids).exclude(is_standard=True))
            all_v = ([bki] if bki else []) + selected

            intake_map = {}
            for vi in VesselVoyageIntake.objects.filter(
                vessel__in=all_v, voyage__in=voyages_qs
            ).select_related('vessel', 'voyage'):
                intake_map.setdefault(vi.vessel_id, {})[vi.voyage_id] = vi.intake

            voyages = [
                {
                    'name': v.name, 'ballast_dist': v.ballast_dist,
                    'laden_dist': v.laden_dist, 'load_rate': v.load_rate,
                    'dis_rate': v.dis_rate, 'load_factor': v.load_factor,
                    'dis_factor': v.dis_factor,
                    'turntimes_hours': v.turntimes_hours,
                    'port_exp': v.port_exp, 'various_exp': v.various_exp,
                }
                for v in voyages_qs
            ]
            vessels_input = []
            for v in all_v:
                v_intakes = intake_map.get(v.id, {})
                vessels_input.append({
                    'name': v.name,
                    'intakes': [v_intakes.get(voy.id, v.default_intake)
                                for voy in voyages_qs],
                    'laden_speed': v.laden_speed,
                    'ballast_speed': v.ballast_speed,
                    'laden_cons': v.laden_cons,
                    'ballast_cons': v.ballast_cons,
                    'port_cons': v.port_cons,
                })

            if voyages and vessels_input and cfg:
                vc = calculate_vessel_comparison(
                    {
                        'hire': cfg.hire, 'ifo_price': cfg.ifo_price,
                        'mgo_price': cfg.mgo_price,
                        'weather_factor': cfg.weather_factor,
                    },
                    voyages,
                    vessels_input,
                )
                blended = float(result['blended_offer'])
                for i, (name, wa) in enumerate(
                    zip(vc['vessels'], vc['weighted_avgs'])
                ):
                    vessel_obj = all_v[i]
                    if vessel_obj.id not in vessel_ids:
                        continue
                    if wa is None:
                        continue
                    vessels_out.append({
                        'id': vessel_obj.id,
                        'name': name,
                        'pct': round(wa * 100, 2),
                        'adjusted_rate': round(blended * wa, 2),
                    })
        except Exception:
            pass  # vessel calc is best-effort; don't let it break the main result

    return JsonResponse({
        'blended_offer': (float(result['blended_offer'])
                          if result['blended_offer'] is not None else None),
        'delivery_date': delivery_date.isoformat(),
        'end_date': result['end_date'].isoformat(),
        'coverage_warning': result.get('coverage_warning'),
        'timeline': timeline,
        'breakdown': [
            {
                'label': r['label'], 'period_type': r.get('period_type', ''),
                'bid': float(r['bid']), 'offer': float(r['offer']),
                'days': r['days'], 'weight': r['weight'],
                'contribution': round(float(r['offer']) * r['weight'], 2),
            }
            for r in result.get('breakdown', [])
        ],
        'vessels': vessels_out,
    })
