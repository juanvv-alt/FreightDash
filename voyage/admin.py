import json
import os
import tempfile
import uuid
from datetime import datetime
from io import BytesIO

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.db.models import Max, Q
from django.utils.html import format_html
from django.shortcuts import redirect, render
from django.urls import path, reverse
from openpyxl import load_workbook

from .models import (
    AvailableIndex,
    CustomIndexPreset,
    DailyIndexValue,
    FreightVoyage,
    RouteParameters,
    VesselFuelConsumption,
    VesselFuelProfile,
    VesselProfile,
    VesselSpeedProfile,
    VoyageFuelSplit,
)


class IndexUploadForm(forms.Form):
    vessel_size = forms.ChoiceField(
        choices=[
            ('capesize', 'Capesize'),
            ('panamax', 'Panamax'),
            ('supramax', 'Supramax'),
            ('handysize', 'Handysize'),
            ('bunker', 'Bunker'),
        ],
        help_text='Assign new index headers to this vessel size if they are not already configured.',
    )
    upload_file = forms.FileField(help_text='Upload preset Excel file (.xlsx) with Date and index columns.')


VESSEL_SIZE_CHOICES = [
    ('capesize', 'Capesize'),
    ('panamax', 'Panamax'),
    ('supramax', 'Supramax'),
    ('handysize', 'Handysize'),
    ('bunker', 'Bunker'),
]


@admin.register(RouteParameters)
class RouteParametersAdmin(admin.ModelAdmin):
    list_display = ('route', 'intake', 'ballast_distance', 'laden_distance', 'updated_at')
    search_fields = ('route',)
    fieldsets = (
        ('Route Information', {
            'fields': ('route',)
        }),
        ('Distance Parameters', {
            'fields': ('ballast_distance', 'laden_distance'),
        }),
        ('Cargo Parameters', {
            'fields': ('intake', 'load_rate', 'discharge_rate', 'turntime_hours'),
        }),
        ('Port Expenses', {
            'fields': ('port_exp_load_port', 'port_exp_discharge_port'),
        }),
        ('Commission & Margins', {
            'fields': ('freight_commission_pct', 'sea_margin_pct'),
        }),
        ('Speed Parameters', {
            'fields': ('ballast_speed', 'laden_speed'),
        }),
        ('Consumption Parameters', {
            'fields': ('ballast_consumption', 'laden_consumption', 'port_consumption'),
        }),
    )
    readonly_fields = ('created_at', 'updated_at')


@admin.register(CustomIndexPreset)
class CustomIndexPresetAdmin(admin.ModelAdmin):
    list_display = ('name', 'updated_at', 'created_at')
    search_fields = ('name',)
    readonly_fields = ('created_at', 'updated_at')


@admin.register(AvailableIndex)
class AvailableIndexAdmin(admin.ModelAdmin):
    list_display = ('name', 'vessel_size', 'order', 'is_active', 'updated_at', 'remove_link')
    list_editable = ('order', 'is_active')
    list_filter = ('vessel_size', 'is_active')
    search_fields = ('name',)
    ordering = ('vessel_size', 'order', 'name')
    readonly_fields = ('created_at', 'updated_at')
    actions = ('remove_selected_indices',)

    @admin.action(description='Remove selected indices')
    def remove_selected_indices(self, request, queryset):
        removed = queryset.count()
        queryset.delete()
        self.message_user(request, f'Removed {removed} index record(s).', level=messages.SUCCESS)

    def remove_link(self, obj):
        url = reverse('admin:voyage_availableindex_delete', args=[obj.pk])
        return format_html('<a class="deletelink" href="{}">Remove</a>', url)

    remove_link.short_description = 'Remove'


@admin.register(DailyIndexValue)
class DailyIndexValueAdmin(admin.ModelAdmin):
    list_display = ('date', 'index', 'value', 'updated_at')
    list_filter = ('index__vessel_size', 'index', 'date')
    search_fields = ('index__name',)
    ordering = ('-date', 'index__name')
    readonly_fields = ('created_at', 'updated_at')


class VesselSpeedProfileInline(admin.TabularInline):
    model = VesselSpeedProfile
    extra = 1
    fields = ('name', 'ballast_speed', 'laden_speed', 'is_default')


class VesselFuelConsumptionInline(admin.TabularInline):
    model = VesselFuelConsumption
    extra = 1
    fields = ('fuel_type', 'sea_consumption', 'port_consumption')


class VesselFuelProfileInline(admin.StackedInline):
    model = VesselFuelProfile
    extra = 1
    fields = ('name', 'is_default')


@admin.register(VesselProfile)
class VesselProfileAdmin(admin.ModelAdmin):
    list_display = ('name', 'vessel_size', 'dwt', 'draft', 'grain_capacity', 'is_active')
    list_filter = ('vessel_size', 'is_active')
    search_fields = ('name',)
    inlines = (VesselSpeedProfileInline, VesselFuelProfileInline)
    readonly_fields = ('created_at', 'updated_at')


@admin.register(VesselFuelProfile)
class VesselFuelProfileAdmin(admin.ModelAdmin):
    list_display = ('name', 'vessel', 'is_default', 'updated_at')
    list_filter = ('is_default', 'vessel__vessel_size')
    search_fields = ('name', 'vessel__name')
    inlines = (VesselFuelConsumptionInline,)
    readonly_fields = ('created_at', 'updated_at')


class VoyageFuelSplitInline(admin.TabularInline):
    model = VoyageFuelSplit
    extra = 1
    fields = ('fuel_index', 'weight_pct')


@admin.register(FreightVoyage)
class FreightVoyageAdmin(admin.ModelAdmin):
    list_display = ('name', 'vessel', 'load_rate', 'discharge_rate', 'ballast_distance', 'laden_distance', 'is_active')
    list_filter = ('is_active', 'vessel__vessel_size', 'intake_mode')
    search_fields = ('name', 'commodity', 'ballast_port')
    inlines = (VoyageFuelSplitInline,)
    fieldsets = (
        ('Voyage Identification', {
            'fields': ('name', 'commodity', 'is_active'),
        }),
        ('Ports', {
            'fields': ('load_ports', 'discharge_ports', 'ballast_port'),
        }),
        ('Port Operations', {
            'fields': (
                'load_rate',
                'discharge_rate',
                'turntime_load_hours',
                'turntime_discharge_hours',
            ),
        }),
        ('Costs and Distances', {
            'fields': (
                'port_exp_load_port',
                'port_exp_discharge_port',
                'misc_expenses',
                'ballast_distance',
                'laden_distance',
            ),
        }),
        ('Vessel and Profiles', {
            'fields': ('vessel', 'speed_profile', 'fuel_profile'),
        }),
        ('Intake', {
            'fields': ('intake_mode', 'intake_manual', 'draft_limit', 'stowage_factor'),
        }),
        ('Margins and Commissions', {
            'fields': (
                'apply_same_sea_margin',
                'sea_margin_ballast_pct',
                'sea_margin_laden_pct',
                'address_commission_pct',
                'brokerage_commission_pct',
            ),
        }),
        ('Daily Hire Link', {
            'fields': ('daily_hire_index',),
        }),
    )
    readonly_fields = ('created_at', 'updated_at')


def _parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _normalize_header(value):
    if value is None:
        return ''
    return str(value).strip().lower()


def _find_excel_header_row(sheet):
    headers = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        cleaned = [str(cell).strip() if cell is not None else '' for cell in row]
        lowered = [value.lower() for value in cleaned]
        if any(name in lowered for name in ['ratedate', 'rate date', 'date']) and any(
            name in lowered for name in ['rateperioddescription', 'rate period description', 'rate period', 'period description', 'period']
        ) and any(name in lowered for name in ['value', 'ratevalue', 'rate value']):
            return row_idx, cleaned
    return None, []


def _find_excel_sheet(workbook, expected_sheet_name='Baltic'):
    normalized_expected = expected_sheet_name.strip().lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == normalized_expected:
            return workbook[sheet_name]
    return None


def upload_indices_view(request):
    if request.method == 'POST':
        form = IndexUploadForm(request.POST, request.FILES)
        if form.is_valid():
            vessel_size = form.cleaned_data['vessel_size']
            upload_file = form.cleaned_data['upload_file']

            try:
                workbook = load_workbook(filename=BytesIO(upload_file.read()), data_only=True, read_only=True)
                sheet = _find_excel_sheet(workbook, expected_sheet_name='Baltic')
                if sheet is None:
                    messages.error(request, "Sheet named ' Baltic' or 'Baltic' was not found in the uploaded file.")
                    return redirect(reverse('admin:indices-upload'))

                header_row_idx, headers = _find_excel_header_row(sheet)
                if not header_row_idx:
                    messages.error(request, 'Header row with RateDate, RatePeriodDescription and Value columns was not found.')
                    return redirect(reverse('admin:indices-upload'))

                normalized_headers = [h.strip() if h else '' for h in headers]
                date_col_idx = None
                period_col_idx = None
                value_col_idx = None
                index_name_col_idx = None

                index_name_candidates = [
                    'indexname', 'index name', 'ratename', 'rate name', 'ratedescription',
                    'rate description', 'index', 'name', 'rate', 'curve', 'curve name', 'rate index'
                ]

                for idx, header in enumerate(normalized_headers):
                    lower_header = header.lower()
                    if lower_header in ['ratedate', 'rate date', 'date']:
                        date_col_idx = idx
                    elif lower_header in ['rateperioddescription', 'rate period description', 'rate period', 'period description', 'period']:
                        period_col_idx = idx
                    elif lower_header in ['value', 'ratevalue', 'rate value', 'rate']:
                        value_col_idx = idx
                    elif any(candidate == lower_header for candidate in index_name_candidates):
                        index_name_col_idx = idx

                if date_col_idx is None:
                    messages.error(request, 'Could not find the RateDate column in the file.')
                    return redirect(reverse('admin:indices-upload'))
                if period_col_idx is None:
                    messages.error(request, 'Could not find the RatePeriodDescription column in the file.')
                    return redirect(reverse('admin:indices-upload'))
                if value_col_idx is None:
                    messages.error(request, 'Could not find the Value column in the file.')
                    return redirect(reverse('admin:indices-upload'))

                if index_name_col_idx is None:
                    for idx, header in enumerate(normalized_headers):
                        if idx not in {date_col_idx, period_col_idx, value_col_idx} and header:
                            index_name_col_idx = idx
                            break

                if index_name_col_idx is None:
                    messages.error(request, 'Could not find an index name column in the file.')
                    return redirect(reverse('admin:indices-upload'))

                extracted_rows = []
                unique_indices = set()
                for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not row:
                        continue

                    rate_period = str(row[period_col_idx]).strip() if len(row) > period_col_idx and row[period_col_idx] is not None else ''
                    if rate_period.lower() != 'spot':
                        continue

                    raw_date = row[date_col_idx] if len(row) > date_col_idx else None
                    parsed_date = _parse_excel_date(raw_date)
                    if not parsed_date:
                        continue

                    raw_value = row[value_col_idx] if len(row) > value_col_idx else None
                    if raw_value in (None, ''):
                        continue
                    try:
                        numeric_value = float(raw_value)
                    except (TypeError, ValueError):
                        continue

                    raw_index_name = row[index_name_col_idx] if len(row) > index_name_col_idx else None
                    index_name = str(raw_index_name).strip() if raw_index_name is not None else ''
                    if not index_name:
                        continue

                    unique_indices.add(index_name)
                    extracted_rows.append({
                        'index_name': index_name,
                        'rate_date': parsed_date.isoformat(),
                        'rate_date_text': parsed_date.strftime('%Y-%m-%d'),
                        'value': numeric_value,
                        'original_period': rate_period,
                    })

                if not extracted_rows:
                    messages.error(request, 'No spot rows were found in the Baltic sheet with valid RateDate and Value entries.')
                    return redirect(reverse('admin:indices-upload'))

                session_id = str(uuid.uuid4())
                temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
                os.makedirs(temp_dir, exist_ok=True)
                temp_file = os.path.join(temp_dir, f'{session_id}.json')

                session_data = {
                    'file_name': upload_file.name,
                    'vessel_size': vessel_size,
                    'sheet_name': sheet.title,
                    'header_row_idx': header_row_idx,
                    'columns': {
                        'date': date_col_idx,
                        'period': period_col_idx,
                        'value': value_col_idx,
                        'index_name': index_name_col_idx,
                    },
                    'rows': extracted_rows,
                    'unique_indices': sorted(unique_indices),
                    'created_at': datetime.now().isoformat(),
                }

                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(session_data, f, indent=2, default=str)

                return redirect(reverse('admin:indices-upload-verify', args=[session_id]))
            except Exception as exc:
                messages.error(request, f'Unable to read or parse the uploaded Excel file: {exc}')
                return redirect(reverse('admin:indices-upload'))
    else:
        form = IndexUploadForm()

    context = {
        **admin.site.each_context(request),
        'title': 'Upload Baltic Indices',
        'form': form,
    }
    return render(request, 'admin/indices_upload.html', context)


def upload_indices_verify_view(request, session_id):
    temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
    temp_file = os.path.join(temp_dir, f'{session_id}.json')

    if not os.path.exists(temp_file):
        messages.error(request, 'Upload verification session not found or has expired.')
        return redirect(reverse('admin:indices-upload'))

    try:
        with open(temp_file, 'r', encoding='utf-8') as f:
            session_data = json.load(f)
    except Exception as exc:
        messages.error(request, f'Unable to load upload session: {exc}')
        return redirect(reverse('admin:indices-upload'))

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'import':
            return _save_excel_indices(request, session_data, temp_file)
        if action == 'discard':
            try:
                os.remove(temp_file)
            except OSError:
                pass
            messages.info(request, 'Excel upload discarded.')
            return redirect(reverse('admin:indices-upload'))

    context = {
        **admin.site.each_context(request),
        'title': 'Review Baltic Index Upload',
        'session_id': session_id,
        'file_name': session_data.get('file_name', ''),
        'vessel_size': session_data.get('vessel_size', ''),
        'sheet_name': session_data.get('sheet_name', ''),
        'rows': session_data.get('rows', []),
        'unique_indices': session_data.get('unique_indices', []),
        'row_count': len(session_data.get('rows', [])),
    }
    return render(request, 'admin/indices_upload_verify.html', context)


def _save_excel_indices(request, session_data, temp_file):
    vessel_size = session_data.get('vessel_size', 'panamax')
    extracted_rows = session_data.get('rows', [])
    if not extracted_rows:
        messages.error(request, 'No extracted rows available to import.')
        return redirect(reverse('admin:indices-upload'))

    unique_names = {row['index_name'].strip() for row in extracted_rows if row.get('index_name')}
    existing_indices = AvailableIndex.objects.filter(name__in=unique_names)
    index_map = {index.name.strip().lower(): index for index in existing_indices}

    if unique_names:
        max_order = AvailableIndex.objects.filter(vessel_size=vessel_size).aggregate(max_order=Max('order'))['max_order'] or 0
    else:
        max_order = 0

    created_indices = 0
    for index_name in sorted(unique_names):
        normalized = index_name.strip().lower()
        if normalized not in index_map:
            max_order += 1
            new_index = AvailableIndex.objects.create(
                name=index_name.strip(),
                vessel_size=vessel_size,
                order=max_order,
                is_active=True,
            )
            index_map[normalized] = new_index
            created_indices += 1

    row_dates = {datetime.fromisoformat(row['rate_date']).date() for row in extracted_rows}
    existing_keys = set(
        DailyIndexValue.objects.filter(
            index__in=index_map.values(),
            date__in=row_dates,
        ).values_list('index_id', 'date')
    )

    values_to_create = []
    skipped_count = 0
    for row in extracted_rows:
        index_name = row['index_name'].strip()
        index_obj = index_map.get(index_name.lower())
        if not index_obj:
            continue

        row_date = datetime.fromisoformat(row['rate_date']).date()
        key = (index_obj.pk, row_date)
        if key in existing_keys:
            skipped_count += 1
            continue

        values_to_create.append(
            DailyIndexValue(
                index=index_obj,
                date=row_date,
                value=row['value'],
            )
        )
        existing_keys.add(key)

    DailyIndexValue.objects.bulk_create(values_to_create, ignore_conflicts=True)
    inserted_count = len(values_to_create)

    try:
        os.remove(temp_file)
    except OSError:
        pass

    messages.success(
        request,
        f'Imported {inserted_count} index values from {len(unique_names)} indices. '
        f'Created {created_indices} new index definitions and skipped {skipped_count} duplicate values.'
    )
    return redirect(reverse('admin:indices-upload'))


def indices_config_view(request):
    valid_vessels = {key for key, _ in VESSEL_SIZE_CHOICES}
    selected_vessel = request.GET.get('vessel') or request.POST.get('vessel') or 'capesize'
    if selected_vessel not in valid_vessels:
        selected_vessel = 'capesize'

    if request.method == 'POST':
        selected_indices_raw = request.POST.getlist('indices')
        selected_indices = []
        seen = set()
        for index_name in selected_indices_raw:
            cleaned = (index_name or '').strip()
            if not cleaned or len(cleaned) > 120 or cleaned in seen:
                continue
            seen.add(cleaned)
            selected_indices.append(cleaned)

        with transaction.atomic():
            # Any index not present in the token list is hidden for this vessel tab.
            AvailableIndex.objects.filter(vessel_size=selected_vessel).update(is_active=False)

            for order, index_name in enumerate(selected_indices, start=1):
                index_obj, _ = AvailableIndex.objects.get_or_create(
                    name=index_name,
                    defaults={
                        'vessel_size': selected_vessel,
                        'order': order,
                        'is_active': True,
                    },
                )
                changed = False
                if index_obj.vessel_size != selected_vessel:
                    index_obj.vessel_size = selected_vessel
                    changed = True
                if index_obj.order != order:
                    index_obj.order = order
                    changed = True
                if not index_obj.is_active:
                    index_obj.is_active = True
                    changed = True
                if changed:
                    index_obj.save(update_fields=['vessel_size', 'order', 'is_active', 'updated_at'])

        messages.success(request, 'Indices display configuration updated.')
        return redirect(f"{reverse('admin:indices-config')}?vessel={selected_vessel}")

    active_indices = list(
        AvailableIndex.objects.filter(vessel_size=selected_vessel, is_active=True)
        .order_by('order', 'name')
        .values_list('name', flat=True)
    )
    suggestions = list(AvailableIndex.objects.order_by('name').values_list('name', flat=True))

    context = {
        **admin.site.each_context(request),
        'title': 'Indices Display Config',
        'selected_vessel': selected_vessel,
        'vessel_choices': VESSEL_SIZE_CHOICES,
        'active_indices': active_indices,
        'all_index_suggestions': suggestions,
    }
    return render(request, 'admin/indices_config.html', context)


_voyage_original_get_urls = admin.site.get_urls


def _voyage_admin_urls():
    custom_urls = [
        path(
            'indices-upload/verify/<str:session_id>/',
            admin.site.admin_view(upload_indices_verify_view),
            name='indices-upload-verify',
        ),
        path(
            'indices-upload/',
            admin.site.admin_view(upload_indices_view),
            name='indices-upload',
        ),
        path(
            'indices-config/',
            admin.site.admin_view(indices_config_view),
            name='indices-config',
        ),
    ]
    return custom_urls + _voyage_original_get_urls()


admin.site.get_urls = _voyage_admin_urls
