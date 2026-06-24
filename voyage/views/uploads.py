import json
import os
import tempfile
import uuid
from datetime import datetime
from io import BytesIO

import logging

import pdfplumber
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models import Max
from django.shortcuts import redirect, render
from django.urls import reverse

from ..models import AvailableIndex, DailyIndexValue
from .helpers import _parse_date


logger = logging.getLogger(__name__)


@staff_member_required(login_url='/admin/login/')
def upload_pdf_indices(request):
    vessel_size_choices = [
        ('capesize', 'Capesize'),
        ('panamax', 'Panamax'),
        ('supramax', 'Supramax'),
        ('handysize', 'Handysize'),
        ('bunker', 'Bunker'),
    ]

    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        vessel_size = request.POST.get('vessel_size', 'panamax')
        pages = request.POST.get('pages', 'all').strip() or 'all'

        if not upload_file:
            messages.error(request, 'Please select a PDF file to upload.')
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        try:
            extracted_tables = _extract_pdf_index_tables(
                upload_file.read(), vessel_size, pages
            )
        except Exception as exc:
            messages.error(request, f'Unable to extract PDF indices: {exc}')
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        if not extracted_tables:
            messages.warning(
                request,
                'No index tables were found in the uploaded PDF. Please check the file and try again.',
            )
            return render(
                request,
                'voyage/upload_pdf_indices.html',
                {
                    'vessel_size_choices': vessel_size_choices,
                    'selected_vessel_size': vessel_size,
                    'pages': pages,
                },
            )

        session_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_pdf_upload')
        os.makedirs(temp_dir, exist_ok=True)
        temp_file = os.path.join(temp_dir, f'{session_id}.json')

        session_data = {
            'file_path': upload_file.name,
            'vessel_size': vessel_size,
            'pages': pages,
            'extracted_tables': extracted_tables,
            'extraction_time': datetime.now().isoformat(),
            'session_id': session_id,
        }

        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(session_data, f, indent=2, default=str)

        return redirect('voyage:verify_pdf_indices', session_id=session_id)

    return render(
        request,
        'voyage/upload_pdf_indices.html',
        {
            'vessel_size_choices': vessel_size_choices,
            'selected_vessel_size': 'panamax',
            'pages': 'all',
        },
    )


def _parse_pdf_pages(pages_str):
    pages_str = str(pages_str or 'all').strip().lower()
    if pages_str in ('all', ''):
        return 'all'

    parsed_pages = []
    for part in pages_str.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            start, end = part.split('-', 1)
            parsed_pages.extend(range(int(start.strip()), int(end.strip()) + 1))
        else:
            parsed_pages.append(int(part))

    return sorted(set(parsed_pages))


def _extract_pdf_index_tables(file_bytes, vessel_size, pages):
    page_numbers = _parse_pdf_pages(pages)

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        if page_numbers == 'all':
            pages_to_process = list(pdf.pages)
        else:
            pages_to_process = [
                pdf.pages[i - 1]
                for i in page_numbers
                if 1 <= i <= len(pdf.pages)
            ]

        if not pages_to_process:
            return []

        existing_indices = {
            idx.name.strip().lower(): idx
            for idx in AvailableIndex.objects.filter(
                vessel_size=vessel_size,
                is_active=True,
            )
        }

        # Try to extract date from PDF text (usually in header)
        document_date = None
        for page in pages_to_process:
            page_text = page.extract_text() or ''
            for line in page_text.split('\n')[:20]:  # Check first 20 lines for date
                parsed = _parse_date(line.strip())
                if parsed:
                    document_date = parsed
                    break
            if document_date:
                break

        extracted_tables = []
        for page_index, page in enumerate(pages_to_process, start=1):
            tables = page.extract_tables()
            if not tables:
                continue

            for table_index, table in enumerate(tables):
                if not table or len(table) < 2:
                    continue

                cleaned_table = [
                    [str(cell).strip() if cell is not None else '' for cell in row]
                    for row in table
                    if row is not None
                ]

                header_row_idx, headers = _find_header_row(cleaned_table)
                if header_row_idx is None or not headers:
                    continue

                # Find date column - could be "Date", "Period", or something like "Today"
                date_col_idx = None
                date_col_is_explicit = False
                for idx, header in enumerate(headers):
                    if header and any(
                        keyword in header.lower()
                        for keyword in ['date', 'period', 'month', 'day']
                    ):
                        date_col_idx = idx
                        date_col_is_explicit = True
                        break

                # If no explicit date column, use the first column as the index name
                if date_col_idx is None:
                    date_col_idx = 0

                index_columns = []
                for idx, header in enumerate(headers):
                    if idx == date_col_idx and date_col_is_explicit:
                        continue
                    if not header or header.lower() in ['change', 'change %', 'change%', '%', 'mtd', 'ytd', 'today']:
                        continue
                    normalized = header.strip()
                    if len(normalized) < 2:
                        continue
                    index_columns.append(
                        {
                            'name': normalized,
                            'column_index': idx,
                            'existing_index': normalized.strip().lower() in existing_indices,
                        }
                    )

                if not index_columns:
                    continue

                processed_rows = []
                for row in cleaned_table[header_row_idx + 1 :]:
                    if len(row) < 1:
                        continue

                    # Get date from row or use document date
                    parsed_date = document_date
                    if date_col_is_explicit and len(row) > date_col_idx:
                        date_str = row[date_col_idx].strip()
                        if date_str:
                            parsed = _parse_date(date_str)
                            if parsed:
                                parsed_date = parsed

                    if not parsed_date:
                        continue

                    # Index name is usually in the first column
                    index_name = row[0].strip() if len(row) > 0 else ''
                    if not index_name or len(index_name) < 2:
                        continue

                    indices = {}

                    # Try to get the "Today" value (usually second column if first is name)
                    # or first numeric column if the date is explicit
                    for idx_col in index_columns:
                        col_idx = idx_col['column_index']
                        if col_idx >= len(row):
                            continue
                        raw_value = row[col_idx]
                        if raw_value is None:
                            continue
                        value_text = str(raw_value).strip()
                        if not value_text or value_text.lower() in ['change', 'change %', 'change%', '%', 'mtd', 'ytd']:
                            continue

                        cleaned_value = (
                            value_text.replace(',', '')
                            .replace('$', '')
                            .replace('£', '')
                            .replace('€', '')
                            .strip()
                        )

                        try:
                            value = float(cleaned_value)
                        except ValueError:
                            continue

                        # Use the row's first column as the index name if header was generic
                        final_index_name = idx_col['name']
                        if idx_col['column_index'] == 0:
                            final_index_name = index_name

                        indices[final_index_name] = {
                            'value': value,
                            'original_value': value_text,
                            'existing': final_index_name.strip().lower() in existing_indices,
                        }

                    if indices:
                        processed_rows.append(
                            {
                                'date': parsed_date.isoformat(),
                                'original_date': parsed_date.strftime('%d %b %Y'),
                                'indices': indices,
                            }
                        )

                if not processed_rows:
                    continue

                extracted_tables.append(
                    {
                        'page': page_index,
                        'table_index': table_index,
                        'data': {
                            'headers': headers,
                            'date_column': date_col_idx,
                            'index_columns': index_columns,
                            'data': processed_rows,
                            'total_rows': len(processed_rows),
                        },
                        'raw_table': cleaned_table[:10],
                    }
                )

        return extracted_tables


def _find_header_row(cleaned_table):
    for row_index, row in enumerate(cleaned_table):
        if any(
            cell
            and any(keyword in cell.lower() for keyword in ['date', 'period', 'month', 'day'])
            for cell in row
        ):
            return row_index, row
    return (0, cleaned_table[0]) if cleaned_table else (None, [])


def verify_pdf_indices(request, session_id):
    """
    View for verifying and validating extracted PDF indices before adding to database.
    """
    temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_pdf_upload')
    temp_file = os.path.join(temp_dir, f'{session_id}.json')

    if not os.path.exists(temp_file):
        messages.error(request, 'Verification session not found or expired.')
        return redirect('voyage:upload_pdf_indices')

    try:
        with open(temp_file, 'r') as f:
            session_data = json.load(f)
    except Exception as e:
        messages.error(request, f'Error loading verification data: {e}')
        return redirect('voyage:upload_pdf_indices')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_selected':
            return _save_selected_indices(request, session_data, temp_file)
        elif action == 'discard':
            # Remove temp file
            try:
                os.remove(temp_file)
            except OSError as exc:
                logger.warning("Could not remove temp file %s: %s", temp_file, exc)
            messages.info(request, 'PDF upload cancelled.')
            return redirect('voyage:upload_pdf_indices')

    # Prepare data for template
    extracted_tables = session_data.get('extracted_tables', [])
    vessel_size = session_data.get('vessel_size', 'panamax')

    # Get existing indices for validation
    existing_indices = {
        idx.name.lower(): idx.name
        for idx in AvailableIndex.objects.filter(vessel_size=vessel_size, is_active=True)
    }

    context = {
        'session_id': session_id,
        'file_path': session_data.get('file_path', ''),
        'vessel_size': vessel_size,
        'extraction_time': session_data.get('extraction_time', ''),
        'extracted_tables': extracted_tables,
        'existing_indices': existing_indices,
        'total_tables': len(extracted_tables),
    }

    return render(request, 'voyage/verify_pdf_indices.html', context)


def _save_selected_indices(request, session_data, temp_file):
    """
    Save selected indices from the verification form.
    """
    selected_tables = request.POST.getlist('selected_tables')
    vessel_size = session_data.get('vessel_size', 'panamax')
    extracted_tables = session_data.get('extracted_tables', [])

    if not selected_tables:
        messages.warning(request, 'No tables selected for import.')
        return redirect(request.META.get('HTTP_REFERER', reverse('voyage:upload_pdf_indices')))

    saved_count = 0
    error_count = 0

    try:
        with transaction.atomic():
            for table_idx_str in selected_tables:
                try:
                    table_idx = int(table_idx_str)
                    if table_idx >= len(extracted_tables):
                        continue

                    table_data = extracted_tables[table_idx]

                    # Create missing indices
                    index_columns = table_data['data']['index_columns']
                    for idx_col in index_columns:
                        if not idx_col.get('existing_index'):
                            # Create new index
                            next_order = AvailableIndex.objects.filter(vessel_size=vessel_size).aggregate(
                                Max('order')
                            )['order__max'] or 0
                            AvailableIndex.objects.create(
                                name=idx_col['name'],
                                vessel_size=vessel_size,
                                order=next_order + 1,
                                is_active=True
                            )

                    # Save daily values
                    from ..models import DailyIndexValue
                    values_to_create = []

                    for row_data in table_data['data']['data']:
                        row_date = datetime.fromisoformat(row_data['date']).date()

                        for idx_name, idx_data in row_data['indices'].items():
                            try:
                                index_obj = AvailableIndex.objects.get(
                                    name__iexact=idx_name,
                                    vessel_size=vessel_size
                                )
                                values_to_create.append(
                                    DailyIndexValue(
                                        index=index_obj,
                                        date=row_date,
                                        value=idx_data['value']
                                    )
                                )
                            except AvailableIndex.DoesNotExist:
                                continue

                    # Bulk create (ignore conflicts for existing data)
                    DailyIndexValue.objects.bulk_create(values_to_create, ignore_conflicts=True)
                    saved_count += len(values_to_create)

                except Exception as e:
                    error_count += 1
                    continue

    except Exception as e:
        messages.error(request, f'Error saving indices: {e}')
        return redirect(request.META.get('HTTP_REFERER', reverse('voyage:upload_pdf_indices')))

    # Clean up temp file
    try:
        os.remove(temp_file)
    except OSError as exc:
        logger.warning("Could not remove temp file %s: %s", temp_file, exc)

    if saved_count > 0:
        messages.success(request, f'Successfully imported {saved_count} index values.')
    if error_count > 0:
        messages.warning(request, f'{error_count} tables had errors during import.')

    return redirect('voyage:upload_pdf_indices')


def _rate_code_vessel_size(code):
    c = (code or '').upper().strip()
    if c.startswith('HS'):
        return 'handysize'
    if c.startswith('C') or c.startswith('BCI'):
        return 'capesize'
    if c.startswith('P') or c.startswith('BPI'):
        return 'panamax'
    if c.startswith('S') or c.startswith('BSI'):
        return 'supramax'
    return 'panamax'


def _parse_excel_spot_rows(upload_file):
    """Parse the Baltic tab of an uploaded .xlsx and return (spot_rows, error_msg).

    spot_rows is a list of dicts: {code, description, date (date obj), value (float)}.
    Returns (None, error_msg) on failure.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
    except Exception as exc:
        return None, f'Could not open file: {exc}'

    if 'Baltic' not in wb.sheetnames:
        return None, "No 'Baltic' sheet found in this workbook."

    ws = wb['Baltic']
    spot_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        rate_code, desc, period, rate_date, value = row[0], row[1], row[2], row[3], row[4]
        if period != 'Spot':
            continue
        if not rate_code or value is None:
            continue
        if isinstance(rate_date, datetime):
            rate_date = rate_date.date()
        elif isinstance(rate_date, str):
            rate_date = _parse_date(rate_date)
        if not rate_date:
            continue
        try:
            value = float(value)
        except (TypeError, ValueError):
            continue
        spot_rows.append({
            'code': str(rate_code).strip(),
            'description': str(desc).strip() if desc else '',
            'date': rate_date.isoformat(),
            'value': value,
        })
    if not spot_rows:
        return None, 'No Spot rows found in the Baltic sheet.'
    return spot_rows, None


@staff_member_required(login_url='/admin/login/')
def upload_excel_indices(request):
    """Upload the daily Baltic Exchange End-of-Day .xlsx — parse and redirect to mapping review."""
    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        if not upload_file:
            messages.error(request, 'Please select an .xlsx file.')
            return render(request, 'voyage/upload_excel_indices.html', {})

        spot_rows, err = _parse_excel_spot_rows(upload_file)
        if err:
            messages.error(request, err)
            return render(request, 'voyage/upload_excel_indices.html', {})

        session_id = str(uuid.uuid4())
        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
        os.makedirs(temp_dir, exist_ok=True)
        temp_file = os.path.join(temp_dir, f'{session_id}.json')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump({'filename': upload_file.name, 'rows': spot_rows}, f)

        return redirect('voyage:review_excel_mappings', session_id=session_id)

    return render(request, 'voyage/upload_excel_indices.html', {})


@staff_member_required(login_url='/admin/login/')
def upload_batch_indices(request):
    """Upload a historical columnar indices .xlsx (Date column + one column per index).

    Only reads the header row during the upload request — the full file is saved to
    disk and processed during the confirm step to avoid request timeouts on large files.
    """
    VESSEL_CHOICES = [
        ('panamax', 'Panamax'),
        ('capesize', 'Capesize'),
        ('supramax', 'Supramax'),
        ('handysize', 'Handysize'),
        ('bunker', 'Bunker'),
    ]

    if request.method == 'POST':
        import openpyxl
        upload_file = request.FILES.get('upload_file')
        vessel_size = request.POST.get('vessel_size', 'panamax')
        if vessel_size not in dict(VESSEL_CHOICES):
            vessel_size = 'panamax'

        if not upload_file:
            messages.error(request, 'Please select an .xlsx file.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
        os.makedirs(temp_dir, exist_ok=True)
        session_id = str(uuid.uuid4())

        # Save the raw file to disk — avoid loading all rows in the web request
        xlsx_path = os.path.join(temp_dir, f'{session_id}.xlsx')
        with open(xlsx_path, 'wb') as fh:
            for chunk in upload_file.chunks():
                fh.write(chunk)

        # Read only the header row to get column names
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as exc:
            os.remove(xlsx_path)
            messages.error(request, f'Could not open file: {exc}')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()

        if not header or str(header[0]).strip().lower() != 'date':
            os.remove(xlsx_path)
            messages.error(request, 'Expected first column to be "Date". Check the file format.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        columns = [str(c).strip() for c in header[1:] if c is not None and str(c).strip()]
        if not columns:
            os.remove(xlsx_path)
            messages.error(request, 'No index columns found in header row.')
            return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})

        # Store session metadata — full parsing happens at confirm time
        meta_file = os.path.join(temp_dir, f'{session_id}.json')
        with open(meta_file, 'w', encoding='utf-8') as f:
            json.dump({
                'type': 'batch',
                'filename': upload_file.name,
                'vessel_size': vessel_size,
                'xlsx_path': xlsx_path,
                'columns': columns,
            }, f)

        return redirect('voyage:review_excel_mappings', session_id=session_id)

    return render(request, 'voyage/upload_batch_indices.html', {'vessel_choices': VESSEL_CHOICES})


@staff_member_required(login_url='/admin/login/')
def review_excel_mappings(request, session_id):
    """Show / confirm per-code mapping between Excel RateCodes and AvailableIndex entries.

    Handles two session types:
      'batch' — large historical files saved as .xlsx on disk; parsed at confirm time
               using streaming + batched bulk_create to avoid timeouts.
      'eod'   — small daily EOD files whose rows were pre-parsed into the session JSON.
    """
    import openpyxl
    from ..models import IndexCodeMapping

    temp_dir = os.path.join(tempfile.gettempdir(), 'freightdash_excel_upload')
    temp_file = os.path.join(temp_dir, f'{session_id}.json')

    if not os.path.exists(temp_file):
        messages.error(request, 'Upload session not found or expired. Please upload again.')
        return redirect('voyage:upload_excel_indices')

    with open(temp_file, 'r', encoding='utf-8') as f:
        session_data = json.load(f)

    filename = session_data.get('filename', '')
    session_type = session_data.get('type', 'eod')       # 'batch' or 'eod'
    session_vessel_size = session_data.get('vessel_size') # may be None for EOD

    # Build unique-codes dict for the review table.
    # Batch sessions store only column names (no row data yet).
    # EOD sessions store pre-parsed rows.
    seen = {}  # code -> representative dict {code, description, date, value}
    if session_type == 'batch':
        for col in session_data.get('columns', []):
            if col not in seen:
                seen[col] = {'code': col, 'description': col, 'date': '—', 'value': ''}
    else:
        for r in session_data.get('rows', []):
            code = r['code']
            if code not in seen:
                seen[code] = r

    # Common DB lookups
    all_indices = list(AvailableIndex.objects.filter(is_active=True).order_by('vessel_size', 'name'))
    saved_mappings = {m.rate_code: m for m in IndexCodeMapping.objects.select_related('target_index').all()}
    index_by_name = {idx.name.lower(): idx for idx in all_indices}

    # ------------------------------------------------------------------ POST
    if request.method == 'POST':
        # Step 1: persist the user's mapping choices
        with transaction.atomic():
            for code in seen:
                action = request.POST.get(f'action_{code}', 'auto')
                target_id = request.POST.get(f'target_{code}', '')

                if action == 'skip':
                    IndexCodeMapping.objects.update_or_create(
                        rate_code=code,
                        defaults={'skip': True, 'target_index': None},
                    )
                    continue

                target_index = None
                if target_id:
                    try:
                        target_index = AvailableIndex.objects.get(pk=int(target_id))
                    except (AvailableIndex.DoesNotExist, ValueError):
                        pass

                if target_index is None:
                    vs = session_vessel_size or _rate_code_vessel_size(code)
                    target_index, _ = AvailableIndex.objects.get_or_create(
                        name=code,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )

                IndexCodeMapping.objects.update_or_create(
                    rate_code=code,
                    defaults={'skip': False, 'target_index': target_index},
                )

        # Step 2: import values — different strategy per session type
        mappings = {m.rate_code: m for m in IndexCodeMapping.objects.select_related('target_index').all()}
        dates_seen = set()
        new_values = 0

        if session_type == 'batch':
            # Stream the saved .xlsx row-by-row; bulk_create in batches of 500
            xlsx_path = session_data.get('xlsx_path', '')
            if not os.path.exists(xlsx_path):
                messages.error(request, 'Uploaded file no longer found. Please upload again.')
                return redirect('voyage:upload_batch_indices')

            # Resolve index objects for each column up-front
            col_index_map = {}   # col_name -> AvailableIndex (or None if skipped)
            for col in session_data.get('columns', []):
                m = mappings.get(col)
                if m and m.skip:
                    col_index_map[col] = None
                    continue
                idx_obj = m.target_index if (m and m.target_index) else None
                if idx_obj is None:
                    vs = session_vessel_size or _rate_code_vessel_size(col)
                    idx_obj, _ = AvailableIndex.objects.get_or_create(
                        name=col,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )
                col_index_map[col] = idx_obj

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            header = None
            index_cols = []   # [(col_offset_in_header, col_name), ...]
            batch = []
            BATCH_SIZE = 500

            for row in ws.iter_rows(min_row=1, values_only=True):
                if header is None:
                    header = row
                    for i, cell in enumerate(header[1:], start=1):
                        name = str(cell).strip() if cell else ''
                        if name and name in col_index_map:
                            index_cols.append((i, name))
                    continue

                date_val = row[0] if row else None
                if date_val is None:
                    continue
                if isinstance(date_val, datetime):
                    row_date = date_val.date()
                else:
                    row_date = _parse_date(str(date_val))
                if not row_date:
                    continue

                for col_i, col_name in index_cols:
                    idx_obj = col_index_map.get(col_name)
                    if idx_obj is None:
                        continue
                    val = row[col_i] if col_i < len(row) else None
                    if val is None:
                        continue
                    try:
                        val = float(val)
                    except (TypeError, ValueError):
                        continue
                    batch.append(DailyIndexValue(index=idx_obj, date=row_date, value=val))
                    dates_seen.add(row_date)
                    if len(batch) >= BATCH_SIZE:
                        created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                        new_values += len(created)
                        batch = []

            wb.close()
            if batch:
                created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                new_values += len(created)

            # Clean up both temp files
            for path in (temp_file, xlsx_path):
                try:
                    os.remove(path)
                except OSError:
                    pass

        else:
            # EOD: rows already parsed in session JSON — use bulk_create too
            batch = []
            BATCH_SIZE = 500
            for r in session_data.get('rows', []):
                code = r['code']
                m = mappings.get(code)
                if m and m.skip:
                    continue
                idx_obj = m.target_index if (m and m.target_index) else None
                if idx_obj is None:
                    vs = session_vessel_size or _rate_code_vessel_size(code)
                    idx_obj, _ = AvailableIndex.objects.get_or_create(
                        name=code,
                        defaults={'vessel_size': vs, 'order': 999, 'is_active': True},
                    )
                row_date = _parse_date(r['date'])
                if not row_date:
                    continue
                batch.append(DailyIndexValue(index=idx_obj, date=row_date, value=r['value']))
                dates_seen.add(row_date)
                if len(batch) >= BATCH_SIZE:
                    created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                    new_values += len(created)
                    batch = []
            if batch:
                created = DailyIndexValue.objects.bulk_create(batch, ignore_conflicts=True)
                new_values += len(created)
            try:
                os.remove(temp_file)
            except OSError:
                pass

        if new_values:
            date_summary = f'{min(dates_seen)} → {max(dates_seen)}' if len(dates_seen) > 3 else ', '.join(str(d) for d in sorted(dates_seen))
            messages.success(request, f'Imported {new_values} index values ({date_summary}).')
        else:
            messages.info(request, 'All values already existed — nothing new imported.')

        redirect_url = 'voyage:upload_batch_indices' if session_type == 'batch' else 'voyage:upload_excel_indices'
        return redirect(redirect_url)

    # ------------------------------------------------------------------ GET
    mapping_rows = []
    for code, rep in seen.items():
        saved = saved_mappings.get(code)
        if saved and not saved.skip and saved.target_index:
            suggested = saved.target_index
            status = 'mapped'
        elif saved and saved.skip:
            suggested = None
            status = 'skip'
        elif code.lower() in index_by_name:
            suggested = index_by_name[code.lower()]
            status = 'exact'
        else:
            suggested = None
            status = 'new'

        mapping_rows.append({
            'code': code,
            'description': rep['description'],
            'date': rep['date'],
            'value': rep['value'],
            'suggested': suggested,
            'status': status,
        })

    mapping_rows.sort(key=lambda r: (0 if r['status'] in ('new', 'skip') else 1, r['code']))

    context = {
        'filename': filename,
        'session_id': session_id,
        'session_type': session_type,
        'mapping_rows': mapping_rows,
        'all_indices': all_indices,
        'total': len(mapping_rows),
        'unmatched': sum(1 for r in mapping_rows if r['status'] == 'new'),
    }
    return render(request, 'voyage/review_excel_mappings.html', context)
