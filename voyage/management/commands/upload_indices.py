from datetime import datetime
from io import BytesIO
import os
import sys

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Max, Q
from openpyxl import load_workbook

from voyage.models import AvailableIndex, DailyIndexValue


class Command(BaseCommand):
    help = 'Upload indices from Excel file to database'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Path to the Excel file containing indices data',
        )
        parser.add_argument(
            '--vessel-size',
            type=str,
            default='panamax',
            choices=['capesize', 'panamax', 'supramax', 'handysize', 'bunker'],
            help='Vessel size for new indices (default: panamax)',
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        vessel_size = options['vessel_size']

        if not os.path.exists(file_path):
            raise CommandError(f'File does not exist: {file_path}')

        try:
            workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
            sheet = workbook.active
        except Exception as e:
            raise CommandError(f'Unable to read Excel file: {e}')

        # Find header row
        header_row_idx = None
        headers = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            cleaned_row = [str(cell).strip() if cell is not None else '' for cell in row]
            if any(cell.lower() == 'date' for cell in cleaned_row):
                header_row_idx = row_idx
                headers = cleaned_row
                break

        if not header_row_idx:
            raise CommandError('Header row with a Date column was not found.')

        # Find date column
        date_col_idx = None
        for idx, name in enumerate(headers):
            if name.lower() == 'date':
                date_col_idx = idx
                break

        if date_col_idx is None:
            raise CommandError('Date column is missing in the uploaded file.')

        # Get data columns
        data_columns = []
        for idx, header in enumerate(headers):
            normalized = header.strip()
            if idx == date_col_idx or not normalized:
                continue
            data_columns.append((idx, normalized))

        if not data_columns:
            raise CommandError('No index columns found in the uploaded file.')

        with transaction.atomic():
            # Normalize headers and find/create indices
            normalized_headers = {
                header.strip().lower(): header.strip()
                for _, header in data_columns
                if header and header.strip()
            }
            existing_query = Q()
            for header in normalized_headers.values():
                existing_query |= Q(name__iexact=header)
            existing_indices = AvailableIndex.objects.filter(existing_query) if existing_query else AvailableIndex.objects.none()
            name_to_index = {
                index.name.strip().lower(): index
                for index in existing_indices
            }

            # Create missing indices
            missing_headers = [
                normalized_headers[normalized]
                for normalized in normalized_headers
                if normalized not in name_to_index
            ]

            if missing_headers:
                next_order = AvailableIndex.objects.filter(vessel_size=vessel_size).aggregate(
                    max_order=Max('order')
                )['max_order'] or 0
                new_indices = []
                for header in missing_headers:
                    next_order += 1
                    new_indices.append(
                        AvailableIndex(
                            name=header,
                            vessel_size=vessel_size,
                            order=next_order,
                            is_active=True,
                        )
                    )
                AvailableIndex.objects.bulk_create(new_indices)
                for index in new_indices:
                    name_to_index[index.name.strip().lower()] = index
                self.stdout.write(
                    self.style.SUCCESS(f'Created {len(new_indices)} new indices: {", ".join(missing_headers)}')
                )

            # Filter data columns
            filtered_data_columns = []
            ignored_columns = []
            for col_idx, index_name in data_columns:
                normalized_name = index_name.strip().lower()
                index_obj = name_to_index.get(normalized_name)
                if index_obj:
                    filtered_data_columns.append((col_idx, index_obj))
                else:
                    ignored_columns.append(index_name.strip())

            if not filtered_data_columns:
                raise CommandError('None of the uploaded index columns exist in Available Indices.')

            # Collect all data
            row_dates = set()
            candidate_values = []
            candidate_keys = set()
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row:
                    continue
                row_date = self._parse_excel_date(row[date_col_idx] if len(row) > date_col_idx else None)
                if not row_date:
                    continue

                row_dates.add(row_date)
                for col_idx, index_obj in filtered_data_columns:
                    raw_value = row[col_idx] if len(row) > col_idx else None
                    if raw_value in (None, ''):
                        continue
                    try:
                        numeric_value = float(raw_value)
                    except (TypeError, ValueError):
                        continue

                    key = (index_obj.pk, row_date)
                    if key in candidate_keys:
                        continue
                    candidate_keys.add(key)
                    candidate_values.append(
                        DailyIndexValue(
                            index=index_obj,
                            date=row_date,
                            value=numeric_value,
                        )
                    )

            # Check existing values
            existing_keys = set(
                DailyIndexValue.objects.filter(
                    index__in=[idx for _, idx in filtered_data_columns],
                    date__in=row_dates,
                ).values_list('index_id', 'date')
            ) if row_dates else set()

            # Filter to create only new values
            values_to_create = []
            for daily_value in candidate_values:
                key = (daily_value.index_id, daily_value.date)
                if key not in existing_keys:
                    values_to_create.append(daily_value)
                    existing_keys.add(key)

            # Bulk create
            DailyIndexValue.objects.bulk_create(values_to_create, ignore_conflicts=True)
            insert_count = len(values_to_create)
            skipped_existing_count = len(candidate_values) - insert_count

            self.stdout.write(
                self.style.SUCCESS(
                    f'Upload completed. Inserted {insert_count} new daily values. '
                    f'Skipped {skipped_existing_count} existing values. '
                    f'Ignored {len(ignored_columns)} unknown columns.'
                )
            )

    def _parse_excel_date(self, value):
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(cleaned, fmt).date()
                except ValueError:
                    continue
        return None